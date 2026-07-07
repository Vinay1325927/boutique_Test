import streamlit as st
from pymongo import MongoClient
import pandas as pd
from datetime import date, datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import os
import time
import hmac
import hashlib
import re
from html import escape as html_escape
import bcrypt
from dotenv import load_dotenv

load_dotenv("credentials/.env")

# =====================================================
# PASSWORD HASH UTILITY
# Run once in a Python shell to generate your hash:
#
#   import bcrypt
#   h = bcrypt.hashpw(b"your_password_here", bcrypt.gensalt())
#   print(h.decode())
#
# Then set PASSWORD_HASH=<output> in credentials/.env
# or in your Streamlit secrets.toml as:
#   PASSWORD_HASH = "<output>"
# =====================================================

# =====================================================
# PAGE CONFIG
# =====================================================

st.set_page_config(
    page_title="Vinay Boutique",
    page_icon="◆",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =====================================================
# CSS — DARK NAVY BLUE THEME
# =====================================================

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,wght@0,300;0,400;0,500;0,600;1,400&family=DM+Serif+Display:ital@0;1&display=swap');

:root {
    /* Light mode palette */
    --bg:           #F4F7FC;
    --bg-2:         #EBF0FA;
    --surface:      #FFFFFF;
    --surface-2:    #F8FAFF;
    --blue:         #2563EB;
    --blue-soft:    #3B82F6;
    --blue-pale:    #BFDBFE;
    --blue-glow:    rgba(37,99,235,0.12);
    --text:         #0F172A;
    --text-2:       #334155;
    --muted:        #64748B;
    --dim:          #94A3B8;
    --emerald:      #059669;
    --rose:         #DC2626;
    --amber:        #D97706;
    --r:            10px;
    --r-lg:         14px;
    --r-xl:         22px;
    --border:       rgba(37,99,235,0.12);
    --border-hover: rgba(37,99,235,0.35);
    --shadow-sm:    0 1px 3px rgba(15,23,42,0.06), 0 1px 2px rgba(15,23,42,0.04);
    --shadow:       0 4px 16px rgba(15,23,42,0.08), 0 2px 6px rgba(15,23,42,0.05);
    --shadow-lg:    0 12px 40px rgba(37,99,235,0.14), 0 4px 12px rgba(15,23,42,0.06);
    /* Legacy aliases kept so existing rules don't break */
    --navy-1:       #F4F7FC;
    --navy-2:       #EBF0FA;
    --navy-3:       #FFFFFF;
    --navy-4:       #F0F5FF;
    --navy-5:       #E2EAFF;
    --cream:        #0F172A;
    --cream-dim:    #334155;
}

/* ━━━ SIDEBAR THEME TOGGLE ━━━ */
[data-testid="stSidebar"] > div:first-child > div:first-child .stButton > button {
    background: transparent !important;
    border: 1px solid var(--border) !important;
    color: var(--muted) !important;
    font-size: 0.7rem !important;
    padding: 0.3rem 0.8rem !important;
    letter-spacing: 0.06em !important;
    width: auto !important;
    float: right;
    margin-bottom: 0.5rem;
}
[data-testid="stSidebar"] > div:first-child > div:first-child .stButton > button:hover {
    border-color: var(--border-hover) !important;
    color: var(--blue) !important;
    transform: none !important;
    background: var(--blue-glow) !important;
}

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif !important;
    background: var(--bg) !important;
    color: var(--text) !important;
}

/* ━━━ ALL INPUTS — DARK TEXT FOR LIGHT MODE ━━━ */
input:not([type="radio"]):not([type="checkbox"]),
textarea,
[data-baseweb="input"] input,
[data-baseweb="base-input"] input,
[data-baseweb="textarea"] textarea,
[data-baseweb="date-picker"] input,
[data-baseweb="select"] input {
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
    caret-color: var(--blue) !important;
    background-color: var(--surface) !important;
}
[data-baseweb="input"] *,
[data-baseweb="base-input"] *,
[data-baseweb="textarea"] * {
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
}
/* Selectbox displayed value */
[data-baseweb="select"] > div > div,
[data-baseweb="select"] > div > div > div,
[data-baseweb="select"] span,
[class*="ValueContainer"] > div,
[class*="singleValue"] {
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
}
/* Number stepper */
.stNumberInput div[data-baseweb="input"] input,
.stNumberInput input[type="number"] {
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
}
/* Date displayed value */
.stDateInput div[data-baseweb="input"] input,
.stDateInput input[type="text"] {
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
}
.stApp {
    background: var(--bg) !important;
    background-image:
        radial-gradient(ellipse 900px 600px at 0% 0%, rgba(37,99,235,0.04) 0%, transparent 70%),
        radial-gradient(ellipse 700px 500px at 100% 100%, rgba(37,99,235,0.03) 0%, transparent 70%);
    background-attachment: fixed;
}
::-webkit-scrollbar { width: 4px; height: 4px; }
::-webkit-scrollbar-track { background: var(--bg-2); border-radius: 99px; }
::-webkit-scrollbar-thumb { background: rgba(37,99,235,0.3); border-radius: 99px; }
::-webkit-scrollbar-thumb:hover { background: rgba(37,99,235,0.5); }

h1, h2, h3 {
    font-family: 'DM Serif Display', serif !important;
    color: var(--text) !important;
    letter-spacing: -0.01em;
    font-weight: 400 !important;
}
h4, h5, h6 {
    font-family: 'DM Sans', sans-serif !important;
    color: var(--muted) !important;
    text-transform: uppercase !important;
    letter-spacing: 0.1em !important;
    font-size: 0.7rem !important;
    font-weight: 600 !important;
}

.page-title {
    font-family: 'DM Serif Display', serif;
    font-size: 2.1rem;
    font-weight: 400;
    color: var(--text);
    letter-spacing: -0.02em;
    line-height: 1.15;
    margin-bottom: 0.2rem;
}
.page-sub {
    font-size: 0.7rem;
    text-transform: uppercase;
    letter-spacing: 0.18em;
    color: var(--dim);
    margin-bottom: 2.4rem;
    font-weight: 500;
}
.rule { height:1px; background:linear-gradient(90deg, var(--blue) 0%, rgba(37,99,235,0.15) 60%, transparent 100%); margin:2rem 0; border:none; }
.rule-sm { height:1px; background:linear-gradient(90deg, rgba(37,99,235,0.25), transparent); margin:1.2rem 0; border:none; }

/* ━━━ SIDEBAR ━━━ */
[data-testid="stSidebar"] {
    background: var(--surface) !important;
    border-right: 1px solid var(--border) !important;
    box-shadow: 2px 0 12px rgba(15,23,42,0.05) !important;
}
[data-testid="stSidebar"] * { color: var(--text) !important; }
.sb-brand { padding: 2rem 1.5rem 0.5rem; text-align: center; }
.sb-logo {
    font-family: 'DM Serif Display', serif;
    font-size: 1.75rem;
    font-weight: 400;
    color: var(--blue);
    letter-spacing: -0.01em;
    line-height: 1;
}
.sb-mark {
    font-size: 0.6rem;
    text-transform: uppercase;
    letter-spacing: 0.22em;
    color: var(--muted);
    margin-top: 0.3rem;
    font-weight: 600;
}
[data-testid="stSidebar"] .stRadio > div {
    gap: 2px !important;
    flex-direction: column !important;
}
[data-testid="stSidebar"] .stRadio > div > label {
    background: transparent !important;
    border: none !important;
    border-radius: var(--r) !important;
    color: var(--muted) !important;
    font-size: 0.84rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.01em !important;
    padding: 0.55rem 1rem 0.55rem 0.75rem !important;
    transition: all 0.18s ease !important;
    cursor: pointer;
    display: flex;
    align-items: center;
}
[data-testid="stSidebar"] .stRadio > div > label:hover {
    background: var(--blue-glow) !important;
    color: var(--blue) !important;
}
[data-testid="stSidebar"] .stRadio > div > label > div:first-child {
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    width: 16px !important;
    height: 16px !important;
    min-width: 16px !important;
    border: 2px solid var(--dim) !important;
    border-radius: 50% !important;
    margin-right: 8px !important;
    background: transparent !important;
    transition: all 0.18s !important;
}
.sb-user {
    font-size: 0.75rem;
    color: var(--muted);
    text-align: center;
    padding: 0.6rem 0 1.5rem;
    letter-spacing: 0.04em;
    font-weight: 500;
}
.sb-sep { height: 1px; background: var(--border); margin: 0.8rem 1rem; }

/* ━━━ PUBLIC BANNER ━━━ */
.pub-banner {
    background: linear-gradient(135deg, #FFFFFF 0%, #F0F6FF 100%);
    border: 1px solid var(--border);
    border-radius: var(--r-xl);
    padding: 2.2rem 2.8rem 1.8rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
    box-shadow: var(--shadow);
}
.pub-banner::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, var(--blue), var(--blue-soft), transparent);
    border-radius: var(--r-xl) var(--r-xl) 0 0;
}
.pub-banner-title {
    font-family: 'DM Serif Display', serif;
    font-size: 2rem;
    font-weight: 400;
    color: var(--text);
    letter-spacing: -0.02em;
    line-height: 1.1;
    margin-bottom: 0.3rem;
}
.pub-banner-sub {
    font-size: 0.68rem;
    text-transform: uppercase;
    letter-spacing: 0.22em;
    color: var(--muted);
    font-weight: 600;
}

/* ━━━ ADMIN LOGIN PANEL (bottom) ━━━ */
.admin-strip {
    margin-top: 3rem;
    border-top: 1px solid var(--border);
    padding-top: 1.5rem;
}
.admin-strip-label {
    font-size: 0.62rem;
    text-transform: uppercase;
    letter-spacing: 0.22em;
    color: var(--dim);
    text-align: center;
    margin-bottom: 0.8rem;
    font-weight: 600;
}

/* ━━━ METRICS ━━━ */
[data-testid="stMetric"] {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--r-lg) !important;
    padding: 1.25rem 1.4rem !important;
    position: relative;
    overflow: hidden;
    box-shadow: var(--shadow-sm) !important;
    transition: box-shadow 0.2s ease, transform 0.2s ease !important;
}
[data-testid="stMetric"]::after {
    content: '';
    position: absolute;
    bottom: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, var(--blue), var(--blue-soft), transparent);
    opacity: 0.6;
}
[data-testid="stMetric"]:hover {
    box-shadow: var(--shadow) !important;
    transform: translateY(-2px);
}
[data-testid="stMetricLabel"] > div {
    color: var(--muted) !important;
    font-size: 0.68rem !important;
    text-transform: uppercase !important;
    letter-spacing: 0.12em !important;
    font-weight: 600 !important;
    font-family: 'DM Sans', sans-serif !important;
}
[data-testid="stMetricValue"] {
    color: var(--text) !important;
    font-family: 'DM Serif Display', serif !important;
    font-size: 1.7rem !important;
    font-weight: 400 !important;
    letter-spacing: -0.02em !important;
    line-height: 1.2 !important;
}

/* ━━━ BUTTONS ━━━ */
.stButton > button {
    background: var(--surface) !important;
    color: var(--blue) !important;
    border: 1.5px solid rgba(37,99,235,0.28) !important;
    border-radius: var(--r) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.8rem !important;
    letter-spacing: 0.04em !important;
    padding: 0.6rem 1.4rem !important;
    transition: all 0.18s ease !important;
    box-shadow: var(--shadow-sm) !important;
    width: 100% !important;
}
.stButton > button:hover {
    background: var(--blue) !important;
    border-color: var(--blue) !important;
    color: #FFFFFF !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 16px rgba(37,99,235,0.25) !important;
}
.stButton > button:active { transform: scale(0.98) !important; }

.stDownloadButton > button {
    background: var(--surface-2) !important;
    color: var(--muted) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--r) !important;
    font-size: 0.78rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.04em !important;
    transition: all 0.18s ease !important;
    width: 100% !important;
}
.stDownloadButton > button:hover {
    border-color: var(--border-hover) !important;
    color: var(--blue) !important;
    background: var(--blue-glow) !important;
}

/* ━━━ FORM SUBMIT ━━━ */
.stForm button[type="submit"] {
    background: var(--blue) !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: var(--r) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.84rem !important;
    letter-spacing: 0.04em !important;
    padding: 0.85rem 2.5rem !important;
    width: 100% !important;
    transition: all 0.2s ease !important;
    box-shadow: 0 2px 8px rgba(37,99,235,0.2) !important;
}
.stForm button[type="submit"]:hover {
    background: #1D4ED8 !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 24px rgba(37,99,235,0.35) !important;
}

/* ━━━ INPUTS ━━━ */
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea > div > div > textarea,
.stDateInput > div > div > input,
.stDateInput input,
input[type="text"], input[type="number"], input[type="date"], textarea {
    background: var(--surface) !important;
    border: 1.5px solid rgba(37,99,235,0.18) !important;
    border-radius: var(--r) !important;
    color: var(--text) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.88rem !important;
    font-weight: 400 !important;
    padding: 0.6rem 0.9rem !important;
    transition: border-color 0.18s ease, box-shadow 0.18s ease !important;
    -webkit-text-fill-color: var(--text) !important;
}
.stTextInput > div > div > input:focus,
.stNumberInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus,
.stDateInput > div > div > input:focus {
    border-color: var(--blue) !important;
    box-shadow: 0 0 0 3px rgba(37,99,235,0.1) !important;
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
}
.stTextInput > div > div > input::placeholder,
.stTextArea > div > div > textarea::placeholder,
.stDateInput > div > div > input::placeholder { color: var(--dim) !important; -webkit-text-fill-color: var(--dim) !important; }

[data-testid="InputInstructions"] {
    display: none !important;
}

/* Date picker calendar popup */
[data-baseweb="calendar"] {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    box-shadow: var(--shadow-lg) !important;
    border-radius: var(--r-lg) !important;
}
[data-baseweb="calendar"] * { color: var(--text) !important; background: transparent !important; }
[data-baseweb="calendar"] [aria-selected="true"] { background: var(--blue) !important; color: #FFFFFF !important; }
[data-baseweb="calendar"] button:hover { background: var(--bg-2) !important; }

/* Date input wrapper */
.stDateInput > div {
    background: var(--surface) !important;
    border-radius: var(--r) !important;
}
.stDateInput > div > div { background: var(--surface) !important; }
.stDateInput svg { fill: var(--muted) !important; }

/* Selectbox */
.stSelectbox > div > div,
.stSelectbox [data-baseweb="select"] > div {
    background: var(--surface) !important;
    border: 1.5px solid rgba(37,99,235,0.18) !important;
    border-radius: var(--r) !important;
    color: var(--text) !important;
    transition: border-color 0.18s ease !important;
}
.stSelectbox > div > div:hover,
.stSelectbox [data-baseweb="select"] > div:hover { border-color: var(--border-hover) !important; }
.stSelectbox [data-baseweb="select"] span,
.stSelectbox [data-baseweb="select"] div { color: var(--text) !important; }

/* Selectbox dropdown menu */
[data-baseweb="popover"] [data-baseweb="menu"],
[data-baseweb="popover"] ul {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    box-shadow: var(--shadow-lg) !important;
    border-radius: var(--r-lg) !important;
}
[data-baseweb="popover"] li,
[data-baseweb="popover"] [role="option"] {
    background: var(--surface) !important;
    color: var(--text) !important;
}
[data-baseweb="popover"] li:hover,
[data-baseweb="popover"] [role="option"]:hover { background: var(--bg-2) !important; }

/* Number input */
.stNumberInput > div > div { background: var(--surface) !important; border-radius: var(--r) !important; }
.stNumberInput input { color: var(--text) !important; -webkit-text-fill-color: var(--text) !important; }

/* ━━━ LABELS ━━━ */
.stTextInput label, .stNumberInput label, .stSelectbox label,
.stTextArea label, .stDateInput label, .stRadio label, .stCheckbox label,
.stTextInput label p, .stNumberInput label p, .stSelectbox label p,
.stTextArea label p, .stDateInput label p, .stRadio label p,
[data-testid="stWidgetLabel"], [data-testid="stWidgetLabel"] p,
[data-testid="stWidgetLabel"] span {
    color: var(--text-2) !important;
    -webkit-text-fill-color: var(--text-2) !important;
    font-size: 0.75rem !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.1em !important;
    font-family: 'DM Sans', sans-serif !important;
}

/* ━━━ DATAFRAME ━━━ */
.stDataFrame {
    border-radius: var(--r-lg) !important;
    border: 1px solid var(--border) !important;
    overflow: hidden !important;
    box-shadow: var(--shadow-sm) !important;
}
[data-testid="stDataFrame"] th {
    background: var(--bg-2) !important;
    color: var(--muted) !important;
    font-size: 0.68rem !important;
    text-transform: uppercase !important;
    letter-spacing: 0.1em !important;
    font-weight: 700 !important;
    border-bottom: 1px solid var(--border) !important;
    padding: 0.75rem 1rem !important;
}
[data-testid="stDataFrame"] td {
    background: var(--surface) !important;
    color: var(--text-2) !important;
    font-size: 0.85rem !important;
    font-weight: 400 !important;
    border-bottom: 1px solid rgba(37,99,235,0.06) !important;
    padding: 0.7rem 1rem !important;
}
[data-testid="stDataFrame"] tr:hover td { background: var(--bg-2) !important; }

/* ━━━ TABS ━━━ */
.stTabs [data-baseweb="tab-list"] {
    background: transparent !important;
    border-bottom: 2px solid var(--border) !important;
    border-radius: 0 !important;
    padding: 0 !important;
    gap: 0 !important;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    border-radius: 0 !important;
    color: var(--dim) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.8rem !important;
    letter-spacing: 0.04em !important;
    padding: 0.75rem 1.25rem !important;
    border-bottom: 2px solid transparent !important;
    transition: all 0.18s !important;
    margin-bottom: -2px !important;
}
.stTabs [aria-selected="true"] {
    color: var(--blue) !important;
    border-bottom-color: var(--blue) !important;
    font-weight: 600 !important;
}
.stTabs [data-baseweb="tab"]:hover { color: var(--text) !important; }

/* ━━━ ALERTS ━━━ */
.stSuccess {
    background: rgba(5,150,105,0.07) !important;
    border: 1px solid rgba(5,150,105,0.25) !important;
    border-radius: var(--r) !important;
}
.stSuccess * { color: #065f46 !important; }
.stInfo {
    background: rgba(37,99,235,0.06) !important;
    border: 1px solid rgba(37,99,235,0.2) !important;
    border-radius: var(--r) !important;
}
.stInfo * { color: #1e3a8a !important; }
.stWarning {
    background: rgba(217,119,6,0.07) !important;
    border: 1px solid rgba(217,119,6,0.25) !important;
    border-radius: var(--r) !important;
}
.stWarning * { color: #92400e !important; }
.stError {
    background: rgba(220,38,38,0.06) !important;
    border: 1px solid rgba(220,38,38,0.22) !important;
    border-radius: var(--r) !important;
}
.stError * { color: #991b1b !important; }

/* ━━━ RADIO (inline) ━━━ */
.stRadio > div { gap: 0.5rem !important; flex-direction: row !important; }
.stRadio > div > label {
    background: var(--surface) !important;
    border: 1.5px solid var(--border) !important;
    border-radius: var(--r) !important;
    padding: 0.5rem 1.1rem !important;
    color: var(--muted) !important;
    font-size: 0.8rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.02em !important;
    transition: all 0.18s !important;
    cursor: pointer;
    box-shadow: var(--shadow-sm) !important;
}
.stRadio > div > label:hover {
    border-color: var(--blue) !important;
    color: var(--blue) !important;
    background: var(--blue-glow) !important;
}

/* ━━━ SECTION HEADERS ━━━ */
.sec-head {
    font-family: 'DM Serif Display', serif;
    font-size: 1.05rem;
    font-weight: 400;
    font-style: italic;
    color: var(--text);
    margin: 1.8rem 0 1rem;
    padding-bottom: 0.6rem;
    border-bottom: 1px solid var(--border);
    letter-spacing: -0.01em;
}

/* ━━━ BADGES ━━━ */
.badge {
    display: inline-block;
    font-size: 0.68rem;
    font-weight: 700;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    padding: 0.22rem 0.65rem;
    border-radius: 6px;
}
.badge-gold  { background: rgba(37,99,235,0.1); color: var(--blue); }
.badge-green { background: rgba(5,150,105,0.1); color: #065f46; }
.badge-red   { background: rgba(220,38,38,0.09); color: #991b1b; }
.badge-muted { background: var(--bg-2); color: var(--muted); }

/* ━━━ PAYMENT COLLECTION ━━━ */
.pay-card {
    background: linear-gradient(135deg, rgba(255,255,255,0.98), rgba(248,250,255,0.96));
    border: 1.5px solid rgba(37,99,235,0.16);
    border-left: 4px solid var(--blue);
    border-radius: var(--r-lg);
    box-shadow: var(--shadow-sm);
    padding: 1rem 1.15rem;
    margin: 0.85rem 0 0.35rem;
}
.pay-grid {
    display: grid;
    grid-template-columns: minmax(240px, 2fr) minmax(150px, 1fr) minmax(140px, 1fr) minmax(140px, 1fr);
    gap: 1rem;
    align-items: center;
}
.pay-name {
    color: var(--text);
    font-size: 0.95rem;
    font-weight: 700;
}
.pay-meta {
    color: var(--muted);
    font-size: 0.78rem;
    margin-top: 0.2rem;
}
.pay-amount {
    color: var(--text);
    font-size: 1rem;
    font-weight: 700;
}
.pay-label {
    color: var(--dim);
    font-size: 0.66rem;
    font-weight: 700;
    letter-spacing: 0.12em;
    text-transform: uppercase;
}
.pay-form-note {
    color: var(--muted);
    font-size: 0.78rem;
    margin: 0.15rem 0 0.75rem;
}
@media (max-width: 760px) {
    .pay-grid { grid-template-columns: 1fr; gap: 0.55rem; }
}

/* ━━━ EMPTY STATE ━━━ */
.empty { text-align: center; padding: 4rem 2rem; color: var(--dim); }
.empty-glyph { font-size: 2rem; margin-bottom: 1rem; color: var(--border); }

/* Number input spinners */
button[data-testid="stNumberInputStepDown"],
button[data-testid="stNumberInputStepUp"] {
    background: var(--bg-2) !important;
    border-color: var(--border) !important;
    color: var(--muted) !important;
}

/* Expanders */
.streamlit-expanderHeader {
    background: var(--surface-2) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--r) !important;
    color: var(--text-2) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.84rem !important;
    font-weight: 500 !important;
}
.streamlit-expanderContent {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-top: none !important;
    border-radius: 0 0 var(--r) var(--r) !important;
    padding: 1rem !important;
}

/* ━━━ BACKUP & RESTORE CARDS ━━━ */
.bk-card {
    background: var(--surface);
    border: 1.5px solid var(--border);
    border-radius: var(--r-xl);
    padding: 1.8rem 2rem;
    box-shadow: var(--shadow);
    transition: box-shadow 0.2s ease, transform 0.2s ease;
    animation: fadeSlideUp 0.4s ease forwards;
    opacity: 0;
}
.bk-card:hover { box-shadow: var(--shadow-lg); transform: translateY(-2px); }
@keyframes fadeSlideUp {
    from { opacity: 0; transform: translateY(16px); }
    to   { opacity: 1; transform: translateY(0);   }
}
.bk-card-icon {
    width: 42px; height: 42px;
    border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-size: 1.25rem;
    margin-bottom: 1rem;
}
.bk-icon-blue  { background: rgba(37,99,235,0.1); }
.bk-icon-green { background: rgba(5,150,105,0.1); }
.bk-icon-amber { background: rgba(217,119,6,0.1); }
.bk-card-title {
    font-family: 'DM Sans', sans-serif;
    font-size: 0.95rem;
    font-weight: 700;
    color: var(--text);
    margin-bottom: 0.2rem;
    letter-spacing: -0.01em;
}
.bk-card-desc {
    font-size: 0.78rem;
    color: var(--muted);
    margin-bottom: 1.2rem;
    line-height: 1.5;
}
.bk-status-badge {
    display: inline-flex; align-items: center; gap: 0.4rem;
    font-size: 0.7rem; font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.1em;
    padding: 0.28rem 0.7rem;
    border-radius: 999px;
}
.bk-status-ok    { background: rgba(5,150,105,0.1); color: #065f46; }
.bk-status-warn  { background: rgba(217,119,6,0.1);  color: #92400e; }
.bk-status-info  { background: rgba(37,99,235,0.1);  color: #1e3a8a; }
.bk-header {
    background: linear-gradient(135deg, #FFFFFF 0%, #EBF3FF 100%);
    border: 1.5px solid var(--border);
    border-radius: var(--r-xl);
    padding: 2rem 2.4rem;
    margin-bottom: 1.8rem;
    position: relative;
    overflow: hidden;
    box-shadow: var(--shadow);
}
.bk-header::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, var(--blue), var(--blue-soft));
    border-radius: var(--r-xl) var(--r-xl) 0 0;
}
.bk-header-icon {
    font-size: 2rem; margin-bottom: 0.6rem; display: block;
}
.bk-header-title {
    font-family: 'DM Serif Display', serif;
    font-size: 1.75rem; font-weight: 400; color: var(--text);
    letter-spacing: -0.02em; line-height: 1.1; margin-bottom: 0.3rem;
}
.bk-header-sub {
    font-size: 0.72rem; text-transform: uppercase; letter-spacing: 0.16em;
    color: var(--muted); font-weight: 600;
}
.bk-progress-bar {
    height: 6px; border-radius: 999px;
    background: linear-gradient(90deg, var(--blue), var(--blue-soft));
    animation: progressPulse 1.5s ease-in-out infinite alternate;
}
@keyframes progressPulse {
    from { opacity: 0.6; width: 30%; }
    to   { opacity: 1;   width: 85%; }
}
.bk-ts {
    font-size: 0.72rem; color: var(--muted);
    font-family: 'DM Sans', sans-serif; margin-top: 0.5rem;
}

/* File uploader */
[data-testid="stFileUploader"] > div {
    background: var(--surface-2) !important;
    border: 2px dashed rgba(37,99,235,0.25) !important;
    border-radius: var(--r-lg) !important;
    transition: border-color 0.18s !important;
}
[data-testid="stFileUploader"] > div:hover {
    border-color: var(--blue) !important;
    background: var(--blue-glow) !important;
}
[data-testid="stFileUploader"] * { color: var(--muted) !important; }

.stCaption, .stCaption * { color: var(--dim) !important; }

p, span { color: var(--text-2); }
</style>
""", unsafe_allow_html=True)

# =====================================================
# PLOTLY TEMPLATE
# =====================================================


# ── THEME TOGGLE — pure session_state, no JS needed ─────────────────────────
if "theme" not in st.session_state:
    st.session_state.theme = "light"

_DARK_CSS = """
<style>
/* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   DARK MODE OVERRIDE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */
:root {
    --bg:           #070C18 !important;
    --bg-2:         #0B1221 !important;
    --surface:      #0F1A2E !important;
    --surface-2:    #152238 !important;
    --blue:         #4D8AE8 !important;
    --blue-soft:    #6BA3F0 !important;
    --blue-pale:    #A8C4F0 !important;
    --blue-glow:    rgba(77,138,232,0.15) !important;
    --text:         #E8EEF8 !important;
    --text-2:       #C8D4E8 !important;
    --muted:        #6A84A8 !important;
    --dim:          #3D5478 !important;
    --border:       rgba(77,138,232,0.18) !important;
    --border-hover: rgba(77,138,232,0.42) !important;
    --shadow-sm:    0 1px 6px rgba(0,0,0,0.4) !important;
    --shadow:       0 4px 24px rgba(0,0,0,0.5) !important;
    --shadow-lg:    0 8px 48px rgba(0,0,0,0.65) !important;
}
.stApp, .stApp > div, .main, [data-testid="stAppViewContainer"], [data-testid="block-container"] {
    background: #070C18 !important;
}
[data-testid="stSidebar"], [data-testid="stSidebar"] > div { background: #0B1221 !important; }
[data-testid="stSidebar"] * { color: #E8EEF8 !important; -webkit-text-fill-color: #E8EEF8 !important; }
p, span, div, label { color: #C8D4E8 !important; }
h1, h2, h3 { color: #E8EEF8 !important; }
[data-testid="stMetric"] { background: #0F1A2E !important; }
[data-testid="stMetricValue"] { color: #E8EEF8 !important; -webkit-text-fill-color: #E8EEF8 !important; }
input, textarea, [data-baseweb="input"] input, [data-baseweb="base-input"] input {
    background: #0F1A2E !important; color: #E8EEF8 !important; -webkit-text-fill-color: #E8EEF8 !important;
}
.stSelectbox > div > div, .stSelectbox [data-baseweb="select"] > div, [data-baseweb="select"] > div { background: #0F1A2E !important; }
[data-baseweb="select"] span, [data-baseweb="select"] div, [class*="singleValue"] { color: #E8EEF8 !important; -webkit-text-fill-color: #E8EEF8 !important; }
[data-baseweb="popover"] [data-baseweb="menu"], [data-baseweb="popover"] ul { background: #0F1A2E !important; }
[data-baseweb="popover"] li, [data-baseweb="popover"] [role="option"] { background: #0F1A2E !important; color: #E8EEF8 !important; }
[data-baseweb="popover"] li:hover, [data-baseweb="popover"] [role="option"]:hover { background: #1C2D47 !important; }
[data-baseweb="calendar"] { background: #0F1A2E !important; }
[data-baseweb="calendar"] * { color: #E8EEF8 !important; }
.stDateInput > div, .stDateInput > div > div { background: #0F1A2E !important; }
[data-testid="stDataFrame"] th { background: #0F1A2E !important; color: #6A84A8 !important; }
[data-testid="stDataFrame"] td { background: #0B1221 !important; color: #C8D4E8 !important; }
[data-testid="stDataFrame"] tr:hover td { background: #0F1A2E !important; }
.stButton > button { background: transparent !important; color: #4D8AE8 !important; border-color: rgba(77,138,232,0.42) !important; }
.stButton > button:hover { background: rgba(77,138,232,0.14) !important; color: #A8C4F0 !important; }
.stDownloadButton > button { background: transparent !important; color: #6A84A8 !important; }
.stTabs [data-baseweb="tab"] { color: #3D5478 !important; }
.stTabs [aria-selected="true"] { color: #E8EEF8 !important; }
.streamlit-expanderHeader { background: #0F1A2E !important; color: #6A84A8 !important; border-color: rgba(77,138,232,0.18) !important; }
.streamlit-expanderContent { background: #0B1221 !important; border-color: rgba(77,138,232,0.18) !important; }
.stRadio > div > label { background: #0F1A2E !important; color: #6A84A8 !important; border-color: rgba(77,138,232,0.18) !important; }
.pub-banner { background: linear-gradient(135deg, #0B1221 0%, #0F1A2E 100%) !important; }
.pub-banner-title { color: #E8EEF8 !important; }
.bk-card { background: #0F1A2E !important; }
.bk-header { background: linear-gradient(135deg, #0B1221 0%, #0F1A2E 100%) !important; }
[data-testid="stWidgetLabel"], [data-testid="stWidgetLabel"] p, [data-testid="stWidgetLabel"] span { color: #C8D4E8 !important; -webkit-text-fill-color: #C8D4E8 !important; }
.stSuccess { background: rgba(61,154,108,0.12) !important; } .stSuccess * { color: #7ADFA0 !important; }
.stInfo { background: rgba(77,138,232,0.1) !important; } .stInfo * { color: #A8C4F0 !important; }
.stWarning { background: rgba(200,160,50,0.1) !important; } .stWarning * { color: #E8C840 !important; }
.stError { background: rgba(192,80,96,0.1) !important; } .stError * { color: #E08090 !important; }
[data-testid="stFileUploader"] > div { background: #0F1A2E !important; border-color: rgba(77,138,232,0.3) !important; }
</style>
"""

def inject_theme():
    if st.session_state.theme == "dark":
        st.markdown(_DARK_CSS, unsafe_allow_html=True)

PLOT_LAYOUT = dict(
    paper_bgcolor="rgba(255,255,255,0)",
    plot_bgcolor="rgba(255,255,255,0)",
    font=dict(family="DM Sans", color="#64748B", size=11),
    title=dict(font=dict(family="DM Serif Display", size=17, color="#0F172A"), pad=dict(b=12), x=0),
    xaxis=dict(gridcolor="rgba(37,99,235,0.08)", linecolor="rgba(37,99,235,0.15)", tickfont=dict(size=10, color="#64748B"), showgrid=True, zeroline=False),
    yaxis=dict(gridcolor="rgba(37,99,235,0.08)", linecolor="rgba(37,99,235,0.15)", tickfont=dict(size=10, color="#64748B"), showgrid=True, zeroline=False),
    legend=dict(bgcolor="rgba(255,255,255,0.92)", bordercolor="rgba(37,99,235,0.15)", borderwidth=1, font=dict(color="#64748B", size=10)),
    margin=dict(l=12, r=12, t=44, b=12),
    colorway=["#2563EB","#3B82F6","#059669","#8BACD8","#DC2626","#1D4ED8","#10B981","#0EA5E9"],
    hoverlabel=dict(bgcolor="rgba(255,255,255,0.97)", bordercolor="rgba(37,99,235,0.3)", font=dict(color="#0F172A", size=11, family="DM Sans"), align="left"),
    bargap=0.35,
)

_DARK_PLOT_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="DM Sans", color="#3D5478", size=11),
    title=dict(font=dict(family="DM Serif Display", size=17, color="#C8D4E8"), pad=dict(b=12), x=0),
    xaxis=dict(gridcolor="rgba(77,138,232,0.08)", linecolor="rgba(77,138,232,0.15)", tickfont=dict(size=10, color="#3D5478"), showgrid=True, zeroline=False),
    yaxis=dict(gridcolor="rgba(77,138,232,0.08)", linecolor="rgba(77,138,232,0.15)", tickfont=dict(size=10, color="#3D5478"), showgrid=True, zeroline=False),
    legend=dict(bgcolor="rgba(7,12,24,0.88)", bordercolor="rgba(77,138,232,0.2)", borderwidth=1, font=dict(color="#6A84A8", size=10)),
    margin=dict(l=12, r=12, t=44, b=12),
    colorway=["#4D8AE8","#6BA3F0","#7ADFA0","#8BACD8","#E08090","#1A3D80","#3D9A6C","#4A9AC8"],
    hoverlabel=dict(bgcolor="rgba(7,12,24,0.96)", bordercolor="rgba(77,138,232,0.3)", font=dict(color="#E8EEF8", size=11, family="DM Sans"), align="left"),
    bargap=0.35,
)

def get_plot_layout():
    """Return plotly layout dict adjusted for current theme."""
    if st.session_state.get("theme") == "dark":
        return _DARK_PLOT_LAYOUT
    return PLOT_LAYOUT

def styled_fig(fig, height=340):
    fig.update_layout(**get_plot_layout(), height=height)
    return fig

# =====================================================
# CONSTANTS
# =====================================================

CATEGORIES      = ["Sarees","Salwar Suits","Lehengas","Kurtis","Western Wear","Accessories","Kids Wear","Blouse","Fabric","Other"]
PAYMENT_METHODS = ["Cash","UPI","Card","Bank Transfer","Part Payment","Credit"]
PAYMENT_COLLECTION_METHODS = ["Cash", "UPI", "Bank Transfer", "Card"]
STATE_OPTIONS   = ["Tamil Nadu","Maharashtra","Karnataka","Delhi","Gujarat","Rajasthan","West Bengal","Uttar Pradesh","Andhra Pradesh","Telangana","Other"]
VENDOR_MANUAL_OPTION = "Add new vendor..."

# =====================================================
# MONGODB
# =====================================================

@st.cache_resource
def get_mongo_client():
    try:
        try:
            uri = st.secrets.get("MONGO_URI", os.getenv("MONGO_URI"))
        except Exception:
            uri = os.getenv("MONGO_URI")
        if not uri:
            st.error("⚠️ MONGO_URI not configured.")
            st.stop()
        client = MongoClient(uri, serverSelectionTimeoutMS=5000)
        client.admin.command("ping")
        return client
    except Exception as e:
        st.error(f"MongoDB connection failed: {e}")
        st.stop()

def get_db():
    return get_mongo_client()["boutique_db"]

def get_col():
    return get_db()["sales"]

def get_next_id():
    counter = get_db()["counters"].find_one_and_update(
        {"_id": "sales_id"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    return counter["seq"]

# =====================================================
# DATA HELPERS
# =====================================================

@st.cache_data(ttl=30)
def fetch_all() -> pd.DataFrame:
    docs = list(get_col().find({}, {"_id": 0}))
    if not docs:
        return pd.DataFrame()
    df = pd.DataFrame(docs)
    for c in ["buying_price", "selling_price", "amount_paid", "pending_amount"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    for c in ["payment_received", "delay_status"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype(int)
    if "sale_date" in df.columns:
        df["sale_date"] = pd.to_datetime(df["sale_date"], errors="coerce")
    df["profit"] = df["selling_price"] - df["buying_price"]
    df["margin"] = (df["profit"] / df["selling_price"].replace(0, 1) * 100).round(2)
    for col in ["vendor", "product_description", "notes", "customer_phone", "last_payment_date", "payment_date", "last_payment_method", "last_payment_received_by"]:
        if col not in df.columns:
            df[col] = ""
    return df

def invalidate_cache():
    fetch_all.clear()
    try:
        get_existing_vendors.clear()
    except NameError:
        pass

def metrics(df: pd.DataFrame) -> dict:
    if df.empty:
        return dict(sales=0, revenue=0, profit=0, pending=0, delayed=0, margin=0, customers=0)
    return dict(
        sales     = len(df),
        revenue   = df["selling_price"].sum(),
        profit    = df["profit"].sum(),
        pending   = df["pending_amount"].sum(),
        delayed   = int((df["delay_status"] == 1).sum()),
        margin    = df["margin"].mean(),
        customers = df["customer_name"].nunique(),
    )

def to_excel(df: pd.DataFrame) -> BytesIO:
    out = BytesIO()
    ex = df.copy()
    if "sale_date" in ex.columns:
        ex["sale_date"] = ex["sale_date"].astype(str)
    ex["profit"]        = (ex["selling_price"] - ex["buying_price"]).round(2)
    ex["profit_margin"] = (ex["profit"] / ex["selling_price"].replace(0, 1) * 100).round(2)
    ex["status"]  = ex["payment_received"].map({0: "Pending", 1: "Received"})
    ex["delayed"] = ex["delay_status"].map({0: "No", 1: "Yes"})
    ordered = ["id","customer_name","customer_phone","sale_date","vendor","product_category","product_description","buying_price","selling_price","profit","profit_margin","amount_paid","pending_amount","status","delayed","payment_method","last_payment_method","last_payment_date","last_payment_received_by","payment_date","notes","created_at"]
    cols = [c for c in ordered if c in ex.columns]
    ex = ex[cols]
    ex.columns = [c.replace("_", " ").title() for c in ex.columns]
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        ex.to_excel(w, index=False)
        ws = w.sheets["Sheet1"]
        for i, col in enumerate(ex.columns, 1):
            ml = max(ex.iloc[:, i-1].astype(str).str.len().max(), len(col)) + 4
            ws.column_dimensions[ws.cell(1, i).column_letter].width = min(ml, 45)
        from openpyxl.styles import Font, PatternFill, Alignment
        blue_fill = PatternFill("solid", fgColor="2E6FD8")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="E8EEF8")
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal="center")
    out.seek(0)
    return out

def get_existing_customers():
    pipeline = [
        {"$match": {"customer_name": {"$ne": None, "$ne": ""}}},
        {"$group": {
            "_id": "$customer_name",
            "visits": {"$sum": 1},
            "last_sale": {"$max": "$sale_date"},
        }},
        {"$sort": {"_id": 1}},
    ]
    return list(get_col().aggregate(pipeline))

def get_existing_customers_with_phone():
    pipeline = [
        {"$match": {"customer_name": {"$ne": None, "$ne": ""}}},
        {"$sort": {"sale_date": -1, "created_at": -1}},
        {"$group": {
            "_id": "$customer_name",
            "phones": {"$push": "$customer_phone"},
            "visits": {"$sum": 1},
            "last_sale": {"$max": "$sale_date"},
        }},
        {"$sort": {"_id": 1}},
    ]
    customers = list(get_col().aggregate(pipeline))
    for customer in customers:
        customer["phone"] = next((str(p).strip() for p in customer.get("phones", []) if str(p or "").strip()), "")
        customer.pop("phones", None)
    return customers

@st.cache_data(ttl=60)
def get_existing_vendors():
    vendors = set()
    for collection_name in ("sales", "inventory"):
        try:
            for vendor in get_db()[collection_name].distinct("vendor"):
                vendor = str(vendor or "").strip()
                if vendor:
                    vendors.add(vendor)
        except Exception:
            pass
    return sorted(vendors, key=str.casefold)

# =====================================================
# HELPERS
# =====================================================

def page_header(title, sub):
    st.markdown(f"<div class='page-title'>{html_escape(title)}</div>", unsafe_allow_html=True)
    st.markdown(f"<div class='page-sub'>{html_escape(sub)}</div>", unsafe_allow_html=True)

def sec(label):
    st.markdown(f"<div class='sec-head'>{html_escape(label)}</div>", unsafe_allow_html=True)

def rule():
    st.markdown("<hr class='rule'>", unsafe_allow_html=True)

def rule_sm():
    st.markdown("<hr class='rule-sm'>", unsafe_allow_html=True)

def is_admin():
    return st.session_state.get("logged_in", False)

def normalize_phone(phone: str) -> str:
    digits = re.sub(r"\D", "", str(phone or ""))
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    return digits[:20]

def first_nonempty(values) -> str:
    for value in values:
        value = str(value or "").strip()
        if value:
            return value
    return ""

def parse_currency(raw: str) -> tuple[float, bool]:
    text = str(raw or "").strip()
    if not text:
        return 0.0, True
    cleaned = re.sub(r"[₹,\s]", "", text)
    if not cleaned:
        return 0.0, True
    try:
        value = float(cleaned)
    except ValueError:
        return 0.0, False
    return value, value >= 0

def money_value(value, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default

def currency_input(label: str, key: str, value: float | None = None) -> tuple[str, float, bool]:
    default = "" if value in (None, 0, 0.0) else f"{float(value):.2f}"
    raw = st.text_input(label, value=default, placeholder="0.00", key=key)
    parsed, valid = parse_currency(raw)
    return raw, parsed, valid

def default_receiver_name() -> str:
    return str(st.session_state.get("username") or "Admin").strip().title()

def record_payment(row: pd.Series, payment_amount: float, payment_date: date, payment_method: str, received_by: str) -> tuple[bool, str]:
    pending = round(max(money_value(row.get("pending_amount")), 0.0), 2)
    amount = round(payment_amount, 2)
    payment_method = str(payment_method or "").strip()
    received_by = str(received_by or "").strip()
    if amount <= 0:
        return False, "Payment amount must be greater than 0."
    if amount > pending:
        return False, "Payment amount cannot exceed the pending amount."
    if not payment_method:
        return False, "Select how the customer paid."
    if not received_by:
        return False, "Enter who received the payment."

    new_pending = round(max(pending - amount, 0.0), 2)
    new_paid = round(money_value(row.get("amount_paid")) + amount, 2)
    payment_entry = {
        "amount": amount,
        "date": str(payment_date),
        "method": payment_method,
        "received_by": received_by[:80],
        "recorded_at": str(datetime.now()),
        "recorded_by": st.session_state.get("username", "Admin"),
    }
    set_fields = {
        "amount_paid": new_paid,
        "pending_amount": new_pending,
        "payment_received": 1 if new_pending == 0 else 0,
        "last_payment_date": str(payment_date),
        "last_payment_method": payment_method,
        "last_payment_received_by": received_by[:80],
        "payment_method": payment_method,
        "updated_at": str(datetime.now()),
    }
    if new_pending == 0:
        set_fields["payment_date"] = str(payment_date)

    get_col().update_one(
        {"id": int(row["id"])},
        {"$set": set_fields, "$push": {"payment_history": payment_entry}},
    )
    invalidate_cache()
    status = "Payment completed." if new_pending == 0 else f"Partial payment saved. ₹{new_pending:,.2f} still pending."
    return True, status

def vendor_picker(label: str, key_prefix: str, current: str = "") -> str:
    current = str(current or "").strip()
    vendors = get_existing_vendors()
    options = [""] + vendors
    if current and current not in options:
        options.insert(1, current)
    options.append(VENDOR_MANUAL_OPTION)
    selected = st.selectbox(
        label,
        options,
        index=options.index(current) if current in options else 0,
        format_func=lambda value: "Select vendor" if value == "" else value,
        key=f"{key_prefix}_select",
    )
    if selected == VENDOR_MANUAL_OPTION:
        return st.text_input("New Vendor", value="" if current in vendors else current, key=f"{key_prefix}_manual").strip()
    return selected.strip()

# =====================================================
# PUBLIC ADD SALE PAGE
# =====================================================

def page_add_sale(public=False):
    if public:
        st.markdown("""
        <div class='pub-banner'>
            <div class='pub-banner-title'>Vinay Boutique</div>
            <div class='pub-banner-sub'>◆ Record a New Sale</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        page_header("New Sale", "Record a Transaction")

    ctype = st.radio("", ["New Customer", "Existing Customer"], horizontal=True)
    rule_sm()

    cname, cphone = "", ""

    if ctype == "Existing Customer":
        if public:
            existing = get_existing_customers()
        else:
            existing = get_existing_customers_with_phone()

        if existing:
            if public:
                opts = [r["_id"] for r in existing]
                sel  = st.selectbox("Select Customer", opts)
                cname = sel
                rec   = next((r for r in existing if r["_id"] == cname), {})
                ca, cb = st.columns(2)
                ca.info(f"**Name:** {cname}")
                cb.info(f"**Visits:** {rec.get('visits', '—')}")
                cphone = ""
            else:
                opts  = [f"{r['_id']}  —  {r.get('phone','') or 'No phone'}" for r in existing]
                sel   = st.selectbox("Select Customer", opts)
                cname = sel.split("  —  ")[0].strip()
                rec   = next((r for r in existing if r["_id"] == cname), {})
                cphone = rec.get("phone", "")
                ca, cb, cc = st.columns(3)
                ca.info(f"**Name:** {cname}")
                cb.info(f"**Phone:** {cphone or 'N/A'}")
                cc.info(f"**Visits:** {rec.get('visits', '—')}")
        else:
            st.warning("No existing customers found.")
            ctype = "New Customer"

    with st.form("sale_form", clear_on_submit=True):
        sec("Customer")
        c1, c2, c3 = st.columns(3)
        with c1:
            cname  = st.text_input("Customer Name *", value=cname,
                                   disabled=(ctype == "Existing Customer"))
        with c2:
            if public and ctype == "Existing Customer":
                cphone = ""
                st.text_input("Phone", value="", placeholder="(Admin access required)", disabled=True)
            else:
                phone_key_seed = "new" if ctype != "Existing Customer" else re.sub(r"[^0-9A-Za-z]+", "_", cname) or "existing"
                cphone = st.text_input("Phone", value=cphone, placeholder="+91 XXXXXXXXXX",
                                       key=f"sale_phone_{phone_key_seed}")
        with c3:
            sdate = st.date_input("Sale Date", date.today())

        sec("Product")
        p1, p2, p3 = st.columns(3)
        with p1: cat  = st.selectbox("Category *", CATEGORIES)
        with p2: vend = vendor_picker("Vendor / Supplier", "sale_vendor")
        with p3: qty  = st.number_input("Quantity", min_value=1, step=1, value=1)
        desc = st.text_area("Description", placeholder="Fabric, colour, design details…", height=70)

        sec("Pricing & Payment")
        pr1, pr2, pr3, pr4 = st.columns(4)
        with pr1: _, buy, buy_ok           = currency_input("Buying Price (₹) *", "sale_buying_price")
        with pr2: _, sell, sell_ok         = currency_input("Selling Price (₹) *", "sale_selling_price")
        with pr3: _, paid_amt, paid_ok     = currency_input("Amount Paid (₹)", "sale_amount_paid")
        with pr4: pm       = st.selectbox("Payment Method", PAYMENT_METHODS)

        pending_amt = max(round(sell - paid_amt, 2), 0.0)
        profit_amt  = round((sell - buy) * qty, 2)
        margin_pct  = round(profit_amt / (sell * qty) * 100, 2) if sell > 0 else 0.0

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Pending",        f"₹{pending_amt:,.2f}")
        m2.metric("Profit (Total)", f"₹{profit_amt:,.2f}")
        m3.metric("Margin",         f"{margin_pct:.1f}%")
        m4.metric("Total Value",    f"₹{sell * qty:,.2f}")

        notes = st.text_area("Notes", placeholder="Special instructions…", height=60)

        submitted = st.form_submit_button("Save Sale", use_container_width=True)

        if submitted:
            # ── Public form rate limiting ────────────────────────────────
            if public:
                now = time.time()
                window_start = st.session_state.get("pub_window_start", now)
                pub_count    = st.session_state.get("pub_submit_count", 0)
                # Reset counter every 60 seconds
                if now - window_start > 60:
                    st.session_state.pub_window_start  = now
                    st.session_state.pub_submit_count  = 0
                    pub_count = 0
                pub_count += 1
                st.session_state.pub_submit_count = pub_count
                if pub_count > 5:
                    st.error("Too many submissions. Please wait a minute before trying again.")
                    st.stop()

            # ── Input length guards ──────────────────────────────────────
            MAX = {"name": 120, "phone": 20, "vendor": 100, "desc": 500, "notes": 500}
            errs = []
            if len(cname) > MAX["name"]:   errs.append(f"Customer name must be under {MAX['name']} characters.")
            if len(cphone) > MAX["phone"]:  errs.append(f"Phone number must be under {MAX['phone']} characters.")
            if len(vend) > MAX["vendor"]:   errs.append(f"Vendor name must be under {MAX['vendor']} characters.")
            if len(desc) > MAX["desc"]:     errs.append(f"Description must be under {MAX['desc']} characters.")
            if len(notes) > MAX["notes"]:   errs.append(f"Notes must be under {MAX['notes']} characters.")

            if not cname.strip():  errs.append("Customer name is required.")
            if not buy_ok:         errs.append("Buying price must be a valid number.")
            elif buy  <= 0:        errs.append("Buying price must be > 0.")
            if not sell_ok:        errs.append("Selling price must be a valid number.")
            elif sell <= 0:        errs.append("Selling price must be > 0.")
            if not paid_ok:        errs.append("Amount paid must be a valid number.")
            elif sell_ok and paid_amt > sell:
                errs.append("Amount paid cannot exceed selling price.")
            if buy_ok and sell_ok and sell < buy:
                st.warning("Selling price is below buying price — this sale will be a loss.")

            if errs:
                for e in errs: st.error(e)
            else:
                get_col().insert_one({
                    "id":                  get_next_id(),
                    "customer_name":       cname.strip()[:120],
                    "customer_phone":      normalize_phone(cphone),
                    "sale_date":           str(sdate),
                    "vendor":              vend.strip()[:100],
                    "product_category":    cat,
                    "product_description": desc.strip()[:500],
                    "quantity":            qty,
                    "buying_price":        round(buy, 2),
                    "selling_price":       round(sell, 2),
                    "amount_paid":         round(paid_amt, 2),
                    "pending_amount":      pending_amt,
                    "payment_received":    1 if pending_amt == 0 else 0,
                    "delay_status":        0,
                    "payment_method":      pm,
                    "notes":               notes.strip()[:500],
                    "created_at":          str(datetime.now()),
                })
                invalidate_cache()
                st.success(f"✓ Sale recorded for {cname.strip()}.")
                st.balloons()
                st.rerun()

# =====================================================
# AUTH HELPERS
# =====================================================

_MAX_ATTEMPTS   = 5
_LOCKOUT_SECS   = 300   # 5-minute lockout after max attempts
_BACKOFF_BASE   = 1.5   # seconds — doubles each attempt after 1st failure

def _get_stored_hash() -> bytes | None:
    """
    Return the bcrypt hash of the admin password.

    Priority:
      1. st.secrets["PASSWORD_HASH"]  — a bcrypt hash (preferred for production)
      2. st.secrets["PASSWORD"]       — plain text, hashed on the fly (migration path)
      3. env var PASSWORD_HASH        — bcrypt hash
      4. env var PASSWORD             — plain text, hashed on the fly

    If none of these are set the function returns None and login is blocked.
    """
    try:
        secrets = st.secrets
    except Exception:
        secrets = {}

    # 1. Pre-hashed secret
    h = secrets.get("PASSWORD_HASH") or os.getenv("PASSWORD_HASH", "")
    if h:
        return h.encode() if isinstance(h, str) else h

    # 2. Plain-text secret — hash on the fly (one-time cost per cold start)
    p = secrets.get("PASSWORD") or os.getenv("PASSWORD", "")
    if p:
        return bcrypt.hashpw(p.encode(), bcrypt.gensalt())

    # Nothing configured — fail closed
    return None


def _get_username() -> str | None:
    try:
        u = st.secrets.get("USERNAME") or os.getenv("USERNAME", "")
    except Exception:
        u = os.getenv("USERNAME", "")
    return u.strip() or None


def _check_lockout() -> tuple[bool, int]:
    """Returns (is_locked, seconds_remaining)."""
    attempts  = st.session_state.get("login_attempts", 0)
    lock_time = st.session_state.get("login_lock_until", 0)
    now       = time.time()
    if lock_time and now < lock_time:
        return True, int(lock_time - now)
    if lock_time and now >= lock_time:
        # Reset after lockout expires
        st.session_state.login_attempts   = 0
        st.session_state.login_lock_until = 0
    return False, 0


def _record_failure():
    attempts = st.session_state.get("login_attempts", 0) + 1
    st.session_state.login_attempts = attempts
    if attempts >= _MAX_ATTEMPTS:
        st.session_state.login_lock_until = time.time() + _LOCKOUT_SECS
    else:
        # Progressive back-off delay (no await needed — this is server-side Streamlit)
        delay = _BACKOFF_BASE * (2 ** (attempts - 1))
        time.sleep(min(delay, 30))


def _verify_credentials(username: str, password: str) -> bool:
    stored_user = _get_username()
    stored_hash = _get_stored_hash()

    if stored_user is None or stored_hash is None:
        return False

    # Constant-time username compare
    user_ok = hmac.compare_digest(username.encode(), stored_user.encode())
    # bcrypt compare (constant-time internally)
    try:
        pass_ok = bcrypt.checkpw(password.encode(), stored_hash)
    except Exception:
        pass_ok = False

    return user_ok and pass_ok


# =====================================================
# ADMIN LOGIN
# =====================================================

def render_admin_login_strip():
    st.markdown("<div class='admin-strip'>", unsafe_allow_html=True)
    st.markdown("<div class='admin-strip-label'>◆ Admin Access</div>", unsafe_allow_html=True)

    # Credential config check (fail closed if nothing set)
    if _get_stored_hash() is None or _get_username() is None:
        with st.expander("Sign in to Admin Dashboard", expanded=False):
            st.error(
                "Admin credentials are not configured. "
                "Set USERNAME and PASSWORD (or PASSWORD_HASH) in st.secrets or environment variables."
            )
        st.markdown("</div>", unsafe_allow_html=True)
        return

    with st.expander("Sign in to Admin Dashboard", expanded=False):
        locked, secs_left = _check_lockout()
        if locked:
            st.error(f"Too many failed attempts. Try again in {secs_left // 60}m {secs_left % 60}s.")
            st.markdown("</div>", unsafe_allow_html=True)
            return

        attempts_left = _MAX_ATTEMPTS - st.session_state.get("login_attempts", 0)
        if attempts_left < _MAX_ATTEMPTS:
            st.warning(f"{attempts_left} attempt(s) remaining before lockout.")

        with st.form("admin_login_form"):
            u = st.text_input("Username", placeholder="username",   key="admin_u")
            p = st.text_input("Password", type="password",
                              placeholder="••••••••",                key="admin_p")
            submitted = st.form_submit_button("Sign In", use_container_width=True)

        if submitted:
            if _verify_credentials(u, p):
                st.session_state.logged_in      = True
                st.session_state.username        = u
                st.session_state.login_attempts  = 0
                st.session_state.login_lock_until = 0
                st.rerun()
            else:
                _record_failure()
                locked2, secs_left2 = _check_lockout()
                if locked2:
                    st.error(f"Account locked for {secs_left2 // 60}m {secs_left2 % 60}s.")
                else:
                    st.error("Invalid credentials.")

    st.markdown("</div>", unsafe_allow_html=True)

# =====================================================
# ADMIN SIDEBAR
# =====================================================

def sidebar():
    with st.sidebar:
        # ── THEME TOGGLE ──────────────────────────────────────────────────
        is_light = st.session_state.theme == "light"
        toggle_label = "🌙 Dark Mode" if is_light else "☀️ Light Mode"
        if st.button(toggle_label, key="theme_toggle"):
            st.session_state.theme = "dark" if is_light else "light"
            st.rerun()

        st.markdown("""
        <div class='sb-brand'>
            <div class='sb-logo' style='font-family:"DM Serif Display",serif'>Vinay</div>
            <div class='sb-mark'>Boutique Manager</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<div class='sb-sep'></div>", unsafe_allow_html=True)

        df = fetch_all()
        m  = metrics(df)

        c1, c2 = st.columns(2)
        c1.metric("Pending", f"₹{m['pending']:,.0f}")
        c2.metric("Profit",  f"₹{m['profit']:,.0f}")
        c1.metric("Sales",   m["sales"])
        c2.metric("Clients", m["customers"])

        st.markdown("<div class='sb-sep'></div>", unsafe_allow_html=True)

        nav = st.radio("Navigation", [
            "Dashboard",
            "Add Sale",
            "Review Accounts",
            "Update Transaction",
            "Customer List",
            "Analytics",
            "Reminders & Alerts",
            "Backup & Restore",
            "Logout",
        ], label_visibility="collapsed")

        st.markdown("<div class='sb-sep'></div>", unsafe_allow_html=True)
        st.markdown(
            f"<div class='sb-user'>◆ {st.session_state.get('username','Admin').title()}</div>",
            unsafe_allow_html=True,
        )
    return nav

# =====================================================
# ADMIN PAGES
# =====================================================

def page_dashboard():
    page_header("Dashboard", "Business Overview")
    df = fetch_all()
    m  = metrics(df)

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Sales",      m["sales"])
    c2.metric("Revenue",    f"₹{m['revenue']:,.0f}")
    c3.metric("Net Profit", f"₹{m['profit']:,.0f}")
    c4.metric("Pending",    f"₹{m['pending']:,.0f}")
    c5.metric("Avg Margin", f"{m['margin']:.1f}%")
    c6.metric("Customers",  m["customers"])

    rule()

    if df.empty:
        st.markdown("<div class='empty'><div class='empty-glyph'>◆</div><div>No sales yet.</div></div>", unsafe_allow_html=True)
        return

    df["month"] = df["sale_date"].dt.to_period("M").astype(str)

    cl, cr = st.columns([3, 2])
    with cl:
        monthly = df.groupby("month").agg(revenue=("selling_price","sum"), profit=("profit","sum"), sales=("id","count")).reset_index()
        fig = go.Figure()
        fig.add_trace(go.Bar(x=monthly["month"], y=monthly["revenue"], name="Revenue", marker_color="rgba(46,111,216,0.4)", marker_line_color="#2E6FD8", marker_line_width=1))
        fig.add_trace(go.Scatter(x=monthly["month"], y=monthly["profit"], name="Profit", mode="lines+markers", line=dict(color="#7ADFA0", width=2), marker=dict(size=5, color="#7ADFA0")))
        styled_fig(fig, 300).update_layout(title="Monthly Revenue & Profit", barmode="overlay", legend=dict(orientation="h", y=1.18, x=0))
        st.plotly_chart(fig, use_container_width=True)

    with cr:
        paid    = (df["payment_received"] == 1).sum()
        pending = (df["payment_received"] == 0).sum()
        fig2 = go.Figure(go.Pie(labels=["Collected","Pending"], values=[paid, pending], hole=0.72, marker=dict(colors=["#2E6FD8","#0F1A2E"]), textfont=dict(size=11), hovertemplate="%{label}: %{value}<extra></extra>"))
        fig2.add_annotation(text=f"<b>{paid+pending}</b>", x=0.5, y=0.52, showarrow=False, font=dict(color="#E8EEF8", family="Playfair Display", size=28))
        fig2.add_annotation(text="sales", x=0.5, y=0.38, showarrow=False, font=dict(color="#3D5478", family="Jost", size=11))
        styled_fig(fig2, 300).update_layout(title="Payment Status", showlegend=True, legend=dict(orientation="h", y=-0.05, x=0.25))
        st.plotly_chart(fig2, use_container_width=True)

    cl2, cr2 = st.columns(2)
    with cl2:
        cat_rev = df.groupby("product_category")["selling_price"].sum().reset_index()
        fig3 = px.pie(cat_rev, values="selling_price", names="product_category", title="Revenue by Category", hole=0.55, color_discrete_sequence=["#2E6FD8","#4D8AE8","#7ADFA0","#8BACD8","#E08090","#1A3D80","#3D9A6C","#4A9AC8","#9B9070","#A8C4F0"])
        styled_fig(fig3, 270); st.plotly_chart(fig3, use_container_width=True)

    with cr2:
        daily = df.set_index("sale_date")["selling_price"].resample("D").sum().reset_index()
        daily.columns = ["date","revenue"]
        daily["rolling"] = daily["revenue"].rolling(7, min_periods=1).mean()
        fig4 = go.Figure()
        fig4.add_trace(go.Bar(x=daily["date"], y=daily["revenue"], name="Daily", marker_color="rgba(46,111,216,0.25)", marker_line_width=0))
        fig4.add_trace(go.Scatter(x=daily["date"], y=daily["rolling"], name="7-day avg", line=dict(color="#2E6FD8", width=1.8)))
        styled_fig(fig4, 270).update_layout(title="Daily Revenue", legend=dict(orientation="h", y=1.18, x=0))
        st.plotly_chart(fig4, use_container_width=True)

    sec("Recent Transactions")
    recent = df.sort_values("sale_date", ascending=False).head(10).copy()
    recent["sale_date"] = recent["sale_date"].dt.strftime("%d %b %Y")
    recent["Status"]    = recent["payment_received"].map({1:"Paid", 0:"Pending"})
    recent["Delayed"]   = recent["delay_status"].map({0:"—", 1:"Yes"})
    show = recent[["id","customer_name","sale_date","product_category","selling_price","profit","pending_amount","Status","Delayed"]].copy()
    show.columns = ["ID","Customer","Date","Category","Amount ₹","Profit ₹","Pending ₹","Status","Delayed"]
    st.dataframe(show, use_container_width=True, hide_index=True)

    rule()
    da, db, _ = st.columns([1, 1, 2])
    with da:
        st.download_button("Export CSV", data=df.assign(sale_date=df["sale_date"].astype(str)).to_csv(index=False), file_name=f"boutique_{date.today()}.csv", mime="text/csv", use_container_width=True)
    with db:
        st.download_button("Export Excel", data=to_excel(df), file_name=f"boutique_{date.today()}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)


def page_review():
    page_header("Accounts", "All Transactions")
    df = fetch_all()
    if df.empty:
        st.markdown("<div class='empty'><div class='empty-glyph'>◆</div><div>No transactions yet.</div></div>", unsafe_allow_html=True)
        return

    with st.expander("Filter & Sort", expanded=True):
        c1, c2, c3, c4 = st.columns(4)
        with c1: srch  = st.text_input("Customer / Phone")
        with c2: catf  = st.selectbox("Category",   ["All"] + CATEGORIES)
        with c3: payf  = st.selectbox("Payment",    ["All","Paid","Pending"])
        with c4: dlayf = st.selectbox("Delay Flag", ["All","On Time","Delayed"])
        c5, c6, c7 = st.columns(3)
        with c5: sortby = st.selectbox("Sort By", ["Date ↓","Date ↑","Amount ↓","Pending ↓","Profit ↓"])
        with c6: d_from = st.date_input("From", value=date.today() - timedelta(days=90))
        with c7: d_to   = st.date_input("To",   value=date.today())

    fdf = df.copy()
    if srch:
        phone_text = fdf["customer_phone"].astype(str)
        mask = (
            fdf["customer_name"].str.contains(srch, case=False, na=False, regex=False)
            | phone_text.str.contains(srch, case=False, na=False, regex=False)
        )
        phone_digits = normalize_phone(srch)
        if phone_digits:
            mask = mask | phone_text.map(normalize_phone).str.contains(phone_digits, na=False, regex=False)
        fdf = fdf[mask]
    if catf  != "All": fdf = fdf[fdf["product_category"] == catf]
    if payf  == "Paid":     fdf = fdf[fdf["payment_received"] == 1]
    elif payf == "Pending": fdf = fdf[fdf["payment_received"] == 0]
    if dlayf == "On Time":  fdf = fdf[fdf["delay_status"] == 0]
    elif dlayf == "Delayed": fdf = fdf[fdf["delay_status"] == 1]
    fdf = fdf[(fdf["sale_date"] >= pd.Timestamp(d_from)) & (fdf["sale_date"] <= pd.Timestamp(d_to))]
    sm = {"Date ↓":("sale_date",False),"Date ↑":("sale_date",True),"Amount ↓":("selling_price",False),"Pending ↓":("pending_amount",False),"Profit ↓":("profit",False)}
    sc, sa = sm[sortby]
    fdf = fdf.sort_values(sc, ascending=sa)

    rule_sm()
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Transactions", len(fdf))
    m2.metric("Revenue",      f"₹{fdf['selling_price'].sum():,.0f}")
    m3.metric("Profit",       f"₹{fdf['profit'].sum():,.0f}")
    m4.metric("Pending",      f"₹{fdf['pending_amount'].sum():,.0f}")
    m5.metric("Avg Margin",   f"{fdf['margin'].mean():.1f}%" if not fdf.empty else "—")
    rule_sm()

    show = fdf[["id","customer_name","customer_phone","sale_date","vendor","product_category","buying_price","selling_price","profit","amount_paid","pending_amount","payment_method","last_payment_method","last_payment_date","last_payment_received_by","delay_status","payment_received"]].copy()
    show["sale_date"]        = show["sale_date"].dt.strftime("%d %b %Y")
    show["vendor"] = show["vendor"].fillna("—").replace("", "—")
    show["last_payment_method"] = show["last_payment_method"].fillna("—").replace("", "—")
    show["last_payment_date"] = show["last_payment_date"].fillna("—").replace("", "—")
    show["last_payment_received_by"] = show["last_payment_received_by"].fillna("—").replace("", "—")
    show["delay_status"]     = show["delay_status"].map({0:"—", 1:"Yes"})
    show["payment_received"] = show["payment_received"].map({0:"Pending", 1:"Paid"})
    show.columns = ["ID","Customer","Phone","Date","Vendor","Category","Buy ₹","Sell ₹","Profit ₹","Paid ₹","Pending ₹","Sale Method","Paid Method","Paid Date","Received By","Delayed","Status"]
    st.dataframe(show, use_container_width=True, hide_index=True)

    dc, de, _ = st.columns([1,1,2])
    with dc:
        st.download_button("Export CSV", data=fdf.assign(sale_date=fdf["sale_date"].astype(str)).to_csv(index=False), file_name=f"accounts_{date.today()}.csv", mime="text/csv", use_container_width=True)
    with de:
        st.download_button("Export Excel", data=to_excel(fdf), file_name=f"accounts_{date.today()}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    sec("Mark Payments")
    pend = fdf[fdf["pending_amount"] > 0].sort_values("pending_amount", ascending=False)
    if pend.empty:
        st.success("All payments received for current filter.")
    else:
        st.markdown(f"<span class='badge badge-gold'>{len(pend)} pending — ₹{pend['pending_amount'].sum():,.0f}</span>", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        for _, row in pend.iterrows():
            row_id = int(row["id"])
            sale_day = row["sale_date"].strftime("%d %b %Y") if pd.notna(row["sale_date"]) else "—"
            pending_value = money_value(row.get("pending_amount"))
            vendor_name = str(row.get("vendor", "") or "").strip() or "—"
            st.markdown(
                f"""
                <div class='pay-card'>
                    <div class='pay-grid'>
                        <div>
                            <div class='pay-name'>{html_escape(str(row['customer_name']))}</div>
                            <div class='pay-meta'>{html_escape(str(row.get('product_category', '—')))} · Sale #{row_id}</div>
                        </div>
                        <div>
                            <div class='pay-label'>Vendor</div>
                            <div class='pay-amount'>{html_escape(vendor_name)}</div>
                        </div>
                        <div>
                            <div class='pay-label'>Pending</div>
                            <div class='pay-amount'>₹{pending_value:,.2f}</div>
                        </div>
                        <div>
                            <div class='pay-label'>Sale Date</div>
                            <div class='pay-amount'>{sale_day}</div>
                        </div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            action_cols = st.columns([4, 1])
            with action_cols[1]:
                if st.button("Mark Paid", key=f"pay_open_{row_id}", use_container_width=True):
                    st.session_state.payment_editor_id = row_id
                    st.rerun()

            if st.session_state.get("payment_editor_id") == row_id:
                st.markdown("<div class='pay-form-note'>Enter the received amount. Use the full pending amount for complete payment, or a smaller amount for partial payment.</div>", unsafe_allow_html=True)
                with st.form(f"payment_form_{row_id}"):
                    pc1, pc2, pc3, pc4 = st.columns([1, 1, 1, 1.2])
                    with pc1:
                        _, payment_amount, payment_ok = currency_input("Amount Received (₹)", f"payment_amount_{row_id}", pending_value)
                    with pc2:
                        payment_date = st.date_input("Paid Date", value=date.today(), key=f"payment_date_{row_id}")
                    with pc3:
                        payment_method = st.selectbox("Paid By", PAYMENT_COLLECTION_METHODS, key=f"payment_method_{row_id}")
                    with pc4:
                        received_by = st.text_input("Received By", value=default_receiver_name(), key=f"payment_received_by_{row_id}")
                    save_col, cancel_col = st.columns(2)
                    with save_col:
                        save_payment = st.form_submit_button("Save Payment", use_container_width=True)
                    with cancel_col:
                        cancel_payment = st.form_submit_button("Cancel", use_container_width=True)

                    if save_payment:
                        if not payment_ok:
                            st.error("Payment amount must be a valid number.")
                        else:
                            ok, message = record_payment(row, payment_amount, payment_date, payment_method, received_by)
                            if ok:
                                st.session_state.payment_editor_id = None
                                st.success(message)
                                st.rerun()
                            else:
                                st.error(message)
                    if cancel_payment:
                        st.session_state.payment_editor_id = None
                        st.rerun()


def page_update():
    page_header("Update", "Edit or Delete a Record")
    c1, c2 = st.columns([2,1])
    with c1: sname = st.text_input("Search by Customer Name / Phone")
    with c2: sid   = st.number_input("Or by Sale ID", min_value=0, step=1)

    if not sname and sid == 0:
        st.info("Enter a customer name, phone, or sale ID to search.")
        return

    if sname:
        search_text = sname.strip()
        phone_digits = normalize_phone(search_text)
        terms = [
            {"customer_name": {"$regex": re.escape(search_text), "$options":"i"}},
            {"customer_phone": {"$regex": re.escape(search_text), "$options":"i"}},
        ]
        if phone_digits and phone_digits != search_text:
            terms.append({"customer_phone": {"$regex": re.escape(phone_digits), "$options":"i"}})
        q = {"$or": terms}
    else:
        q = {"id": int(sid)}
    docs = list(get_col().find(q, {"_id":0}))
    if not docs:
        st.warning("No matching transaction found.")
        return

    df = pd.DataFrame(docs)
    for c in ["buying_price","selling_price","amount_paid","pending_amount"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    preview_cols = [c for c in ["id","customer_name","sale_date","product_category","selling_price","pending_amount","payment_received"] if c in df.columns]
    preview = df[preview_cols].copy()
    if "payment_received" in preview.columns:
        preview["payment_received"] = preview["payment_received"].map({0:"Pending",1:"Paid"})
    st.dataframe(preview, use_container_width=True, hide_index=True)

    sel = st.selectbox("Select ID to Edit", df["id"].tolist(), format_func=lambda x: f"#{x} — {df[df['id']==x]['customer_name'].values[0]}")
    row = df[df["id"] == sel].iloc[0]
    rule_sm()

    with st.form("update_form"):
        sec("Customer & Product")
        c1, c2, c3 = st.columns(3)
        with c1:
            nn = st.text_input("Customer Name", value=str(row.get("customer_name","")))
            np = st.text_input("Phone",          value=str(row.get("customer_phone","")))
        with c2:
            ci  = CATEGORIES.index(row["product_category"]) if row.get("product_category") in CATEGORIES else 0
            nc  = st.selectbox("Category", CATEGORIES, index=ci)
            nv  = vendor_picker("Vendor", f"update_vendor_{sel}", str(row.get("vendor","")))
        with c3:
            try:    existing_date = pd.to_datetime(row.get("sale_date")).date()
            except: existing_date = date.today()
            new_date = st.date_input("Sale Date", value=existing_date)
            nqty = st.number_input("Quantity", min_value=1, step=1, value=int(row.get("quantity",1)))

        ndesc = st.text_area("Description", value=str(row.get("product_description","")), height=60)
        sec("Pricing & Payment")
        pr1, pr2, pr3, pr4 = st.columns(4)
        with pr1: _, nb, nb_ok   = currency_input("Buying Price (₹)", f"update_buying_price_{sel}", float(row["buying_price"]))
        with pr2: _, ns, ns_ok   = currency_input("Selling Price (₹)", f"update_selling_price_{sel}", float(row["selling_price"]))
        with pr3: _, npa, npa_ok = currency_input("Amount Paid (₹)", f"update_amount_paid_{sel}", float(row["amount_paid"]))
        with pr4:
            pi  = PAYMENT_METHODS.index(row["payment_method"]) if row.get("payment_method") in PAYMENT_METHODS else 0
            npm = st.selectbox("Payment Method", PAYMENT_METHODS, index=pi)

        nd     = st.checkbox("Mark as Delayed", value=bool(row.get("delay_status",0)))
        nnotes = st.text_area("Notes", value=str(row.get("notes","")), height=60)

        npend   = max(round(ns - npa, 2), 0.0)
        nprofit = round(ns - nb, 2)
        m1, m2, m3 = st.columns(3)
        m1.metric("Updated Pending", f"₹{npend:,.2f}")
        m2.metric("Updated Profit",  f"₹{nprofit:,.2f}")
        m3.metric("Updated Margin",  f"{(nprofit/ns*100 if ns>0 else 0):.1f}%")

        bu, bd = st.columns(2)
        with bu: upd = st.form_submit_button("Save Changes",       use_container_width=True)
        with bd: dlt = st.form_submit_button("Delete Transaction",  use_container_width=True)

        if upd:
            errs = []
            if not nb_ok:
                errs.append("Buying price must be a valid number.")
            if not ns_ok:
                errs.append("Selling price must be a valid number.")
            if not npa_ok:
                errs.append("Amount paid must be a valid number.")
            if ns_ok and npa_ok and npa > ns:
                errs.append("Amount paid cannot exceed selling price.")
            if errs:
                for err in errs:
                    st.error(err)
            else:
                get_col().update_one({"id": sel}, {"$set": {
                    "customer_name": nn.strip(), "customer_phone": normalize_phone(np),
                    "sale_date": str(new_date), "product_category": nc,
                    "vendor": nv.strip(), "product_description": ndesc.strip(),
                    "quantity": nqty, "buying_price": round(nb,2),
                    "selling_price": round(ns,2), "amount_paid": round(npa,2),
                    "pending_amount": npend, "delay_status": int(nd),
                    "payment_method": npm, "notes": nnotes.strip(),
                    "payment_received": 1 if npend==0 else 0,
                    "updated_at": str(datetime.now()),
                }})
                invalidate_cache(); st.success("Transaction updated."); st.rerun()
        if dlt:
            get_col().delete_one({"id": sel})
            invalidate_cache(); st.success("Transaction deleted."); st.rerun()


def page_customers():
    page_header("Customers", "All Clients")
    df = fetch_all()
    if df.empty:
        st.markdown("<div class='empty'><div class='empty-glyph'>◆</div><div>No customers yet.</div></div>", unsafe_allow_html=True)
        return

    summ = (df.sort_values("sale_date", ascending=False).groupby("customer_name").agg(
        phone=("customer_phone", first_nonempty), transactions=("id","count"),
        spent=("selling_price","sum"), pending=("pending_amount","sum"),
        last_visit=("sale_date","max"), profit=("profit","sum"),
    ).reset_index())
    summ["last_visit"] = pd.to_datetime(summ["last_visit"]).dt.strftime("%d %b %Y")
    summ = summ.sort_values("spent", ascending=False).reset_index(drop=True)
    summ["tier"] = pd.cut(summ["spent"], bins=[0,5000,20000,50000,float("inf")], labels=["Bronze","Silver","Gold","Platinum"])

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Customers", len(summ))
    m2.metric("Avg Spend",       f"₹{summ['spent'].mean():,.0f}")
    m3.metric("With Pending",    len(summ[summ["pending"] > 0]))
    m4.metric("Total Revenue",   f"₹{summ['spent'].sum():,.0f}")

    rule_sm()
    c1, c2 = st.columns([2,1])
    with c1: srch   = st.text_input("Search Customer")
    with c2: tier_f = st.selectbox("Tier", ["All","Bronze","Silver","Gold","Platinum"])

    view = summ.copy()
    if srch:
        phone_digits = normalize_phone(srch)
        mask = (
            view["customer_name"].str.contains(srch, case=False, na=False, regex=False)
            | view["phone"].astype(str).str.contains(srch, case=False, na=False, regex=False)
        )
        if phone_digits:
            mask = mask | view["phone"].astype(str).map(normalize_phone).str.contains(phone_digits, na=False, regex=False)
        view = view[mask]
    if tier_f != "All": view = view[view["tier"] == tier_f]

    disp = view.rename(columns={"customer_name":"Customer","phone":"Phone","transactions":"Visits","spent":"Total Spent ₹","pending":"Pending ₹","last_visit":"Last Visit","profit":"Profit ₹","tier":"Tier"})
    st.dataframe(disp.style.format({"Total Spent ₹":"₹{:,.0f}","Pending ₹":"₹{:,.0f}","Profit ₹":"₹{:,.0f}"}), use_container_width=True, hide_index=True)

    dc, de = st.columns(2)
    with dc:
        st.download_button("Export CSV", data=disp.to_csv(index=False), file_name=f"customers_{date.today()}.csv", mime="text/csv", use_container_width=True)
    with de:
        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as w: disp.to_excel(w, index=False)
        out.seek(0)
        st.download_button("Export Excel", data=out, file_name=f"customers_{date.today()}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    sec("Purchase History")
    chosen = st.selectbox("Select Customer", summ["customer_name"].tolist())
    if chosen:
        hist = df[df["customer_name"] == chosen].sort_values("sale_date", ascending=False).copy()
        hist["status"]   = hist["payment_received"].map({0:"Pending",1:"Paid"})
        hist["sale_date"] = hist["sale_date"].dt.strftime("%d %b %Y")
        h1, h2, h3, h4 = st.columns(4)
        h1.metric("Visits",      len(hist))
        h2.metric("Total Spent", f"₹{hist['selling_price'].sum():,.0f}")
        h3.metric("Pending",     f"₹{hist['pending_amount'].sum():,.0f}")
        h4.metric("Profit",      f"₹{hist['profit'].sum():,.0f}")
        cols = [c for c in ["sale_date","product_category","product_description","selling_price","amount_paid","pending_amount","payment_method","status"] if c in hist.columns]
        show = hist[cols].copy()
        show.columns = ["Date","Category","Description","Price ₹","Paid ₹","Pending ₹","Method","Status"][:len(cols)]
        st.dataframe(show, use_container_width=True, hide_index=True)
        if len(hist) > 1:
            hs = df[df["customer_name"]==chosen].sort_values("sale_date").copy()
            hs["cumulative"] = hs["selling_price"].cumsum()
            fig = px.line(hs, x="sale_date", y="cumulative", title=f"Cumulative Spend — {chosen}", markers=True)
            fig.update_traces(line_color="#2E6FD8", marker_color="#4D8AE8", marker_size=5)
            styled_fig(fig, 230); st.plotly_chart(fig, use_container_width=True)


def page_analytics():
    page_header("Analytics", "Business Intelligence")
    df = fetch_all()
    if df.empty:
        st.info("No data available.")
        return

    df["month"] = df["sale_date"].dt.to_period("M").astype(str)
    df["dow"]   = df["sale_date"].dt.day_name()

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Revenue",       f"₹{df['selling_price'].sum():,.0f}")
    k2.metric("Profit",        f"₹{df['profit'].sum():,.0f}")
    k3.metric("Avg Order",     f"₹{df['selling_price'].mean():,.0f}")
    k4.metric("Avg Margin",    f"{df['margin'].mean():.1f}%")
    k5.metric("Delayed Count", int((df["delay_status"]==1).sum()))

    rule()
    t1, t2, t3, t4, t5 = st.tabs(["Trends","Customers","Categories","Payments","Top Items"])

    with t1:
        c1, c2 = st.columns(2)
        with c1:
            monthly = df.groupby("month").agg(revenue=("selling_price","sum"), profit=("profit","sum")).reset_index()
            fig = go.Figure()
            fig.add_trace(go.Bar(x=monthly["month"], y=monthly["revenue"], name="Revenue", marker_color="rgba(46,111,216,0.4)", marker_line_color="#2E6FD8", marker_line_width=1))
            fig.add_trace(go.Scatter(x=monthly["month"], y=monthly["profit"], name="Profit", mode="lines+markers", line=dict(color="#7ADFA0", width=2), marker=dict(size=5)))
            styled_fig(fig).update_layout(title="Revenue & Profit by Month", barmode="overlay", legend=dict(orientation="h", y=1.18, x=0))
            st.plotly_chart(fig, use_container_width=True)
        with c2:
            daily = df.set_index("sale_date")["selling_price"].resample("D").sum().reset_index()
            daily.columns = ["date","revenue"]
            fig2 = px.area(daily, x="date", y="revenue", title="Daily Revenue")
            fig2.update_traces(fillcolor="rgba(46,111,216,0.12)", line_color="#2E6FD8", line_width=1.5)
            styled_fig(fig2); st.plotly_chart(fig2, use_container_width=True)
        c3, c4 = st.columns(2)
        with c3:
            dow_order = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
            dow = df.groupby("dow").agg(sales=("id","count"), revenue=("selling_price","sum")).reset_index()
            dow["dow"] = pd.Categorical(dow["dow"], categories=dow_order, ordered=True)
            dow = dow.sort_values("dow")
            fig3 = px.bar(dow, x="dow", y="sales", title="Sales by Day of Week", color="revenue", color_continuous_scale=[[0,"#070C18"],[1,"#2E6FD8"]])
            styled_fig(fig3); st.plotly_chart(fig3, use_container_width=True)
        with c4:
            monthly["MoM Growth %"] = monthly["revenue"].pct_change()*100
            fig4 = px.bar(monthly.dropna(), x="month", y="MoM Growth %", title="Month-over-Month Growth", color="MoM Growth %", color_continuous_scale=[[0,"#C05060"],[0.5,"#0F1A2E"],[1,"#7ADFA0"]])
            styled_fig(fig4); st.plotly_chart(fig4, use_container_width=True)

    with t2:
        c1, c2 = st.columns(2)
        with c1:
            top_c = df.groupby("customer_name")["selling_price"].sum().nlargest(10).reset_index()
            fig5 = px.bar(top_c, x="selling_price", y="customer_name", orientation="h", title="Top 10 Customers by Revenue", color="selling_price", color_continuous_scale=[[0,"#070C18"],[1,"#2E6FD8"]])
            styled_fig(fig5); fig5.update_layout(yaxis=dict(autorange="reversed")); st.plotly_chart(fig5, use_container_width=True)
        with c2:
            cp = df.groupby("customer_name")["pending_amount"].sum()
            cp = cp[cp > 0].nlargest(10).reset_index()
            if not cp.empty:
                fig6 = px.bar(cp, x="pending_amount", y="customer_name", orientation="h", title="Top Customers by Pending", color="pending_amount", color_continuous_scale=[[0,"#070C18"],[1,"#C05060"]])
                styled_fig(fig6); fig6.update_layout(yaxis=dict(autorange="reversed")); st.plotly_chart(fig6, use_container_width=True)
            else:
                st.success("No pending amounts.")
        cust_stats = df.groupby("customer_name").agg(visits=("id","count"), revenue=("selling_price","sum"), avg_order=("selling_price","mean")).reset_index()
        fig_scatter = px.scatter(cust_stats, x="visits", y="revenue", size="avg_order", hover_name="customer_name", title="Customer Value Matrix", color="revenue", color_continuous_scale=[[0,"#070C18"],[1,"#2E6FD8"]])
        styled_fig(fig_scatter, 330); st.plotly_chart(fig_scatter, use_container_width=True)
        seg = df.groupby("customer_name").agg(spend=("selling_price","sum")).reset_index()
        seg["tier"] = pd.cut(seg["spend"], bins=[0,5000,20000,50000,float("inf")], labels=["Bronze","Silver","Gold","Platinum"])
        sec("Customer Tier Distribution")
        sg = seg.groupby("tier", observed=True).agg(customers=("customer_name","count"), total=("spend","sum")).reset_index()
        sg.columns = ["Tier","Customers","Total Spend ₹"]
        st.dataframe(sg, use_container_width=True, hide_index=True)

    with t3:
        c1, c2 = st.columns(2)
        with c1:
            cd = df.groupby("product_category").size().reset_index(name="count")
            fig7 = px.pie(cd, values="count", names="product_category", title="Sales Volume by Category", hole=0.55, color_discrete_sequence=["#2E6FD8","#4D8AE8","#7ADFA0","#8BACD8","#E08090","#1A3D80","#3D9A6C","#4A9AC8","#9B9070","#A8C4F0"])
            styled_fig(fig7); st.plotly_chart(fig7, use_container_width=True)
        with c2:
            cp2 = df.groupby("product_category").agg(profit=("profit","sum"), revenue=("selling_price","sum")).reset_index()
            cp2["margin"] = (cp2["profit"]/cp2["revenue"]*100).round(1)
            fig8 = px.bar(cp2, x="product_category", y="profit", title="Profit by Category", color="margin", color_continuous_scale=[[0,"#070C18"],[1,"#7ADFA0"]])
            styled_fig(fig8); st.plotly_chart(fig8, use_container_width=True)
        cm = df.groupby(["month","product_category"])["selling_price"].sum().unstack(fill_value=0)
        if not cm.empty:
            fig9 = px.imshow(cm.T, title="Category × Month Heatmap", color_continuous_scale=[[0,"#070C18"],[0.4,"#1A3D80"],[1,"#2E6FD8"]], aspect="auto")
            styled_fig(fig9, 300); st.plotly_chart(fig9, use_container_width=True)

    with t4:
        c1, c2 = st.columns(2)
        with c1:
            pm = df.groupby("payment_method").size().reset_index(name="count")
            fig10 = px.pie(pm, values="count", names="payment_method", title="Payment Method Distribution", hole=0.58, color_discrete_sequence=["#2E6FD8","#4D8AE8","#7ADFA0","#8BACD8","#1A3D80","#E08090"])
            styled_fig(fig10); st.plotly_chart(fig10, use_container_width=True)
        with c2:
            ps = df.groupby("payment_received").agg(count=("id","count"), total=("pending_amount","sum")).reset_index()
            ps["label"] = ps["payment_received"].map({0:"Pending",1:"Received"})
            fig11 = px.bar(ps, x="label", y="count", title="Payment Status", color="label", color_discrete_map={"Pending":"#2E6FD8","Received":"#7ADFA0"})
            styled_fig(fig11); st.plotly_chart(fig11, use_container_width=True)
        aged = df[df["pending_amount"] > 0].copy()
        if not aged.empty:
            today_ts = pd.Timestamp(date.today())
            aged["days"] = (today_ts - aged["sale_date"]).dt.days
            aged["bucket"] = pd.cut(aged["days"], bins=[0,7,15,30,60,9999], labels=["0–7d","8–15d","16–30d","31–60d","60d+"])
            ag = aged.groupby("bucket", observed=True)["pending_amount"].sum().reset_index()
            fig12 = px.bar(ag, x="bucket", y="pending_amount", title="Pending — Aging Buckets", color="pending_amount", color_continuous_scale=[[0,"#2E6FD8"],[1,"#C05060"]])
            styled_fig(fig12); st.plotly_chart(fig12, use_container_width=True)
        else:
            st.success("No pending payments.")

    with t5:
        c1, c2 = st.columns(2)
        with c1:
            if "vendor" in df.columns:
                vd = (df[df["vendor"].astype(str).str.strip() != ""].groupby("vendor").agg(revenue=("selling_price","sum"), items=("id","count")).nlargest(10,"revenue").reset_index())
                if not vd.empty:
                    fig13 = px.bar(vd, x="revenue", y="vendor", orientation="h", title="Top Vendors by Revenue", color="revenue", color_continuous_scale=[[0,"#070C18"],[1,"#2E6FD8"]])
                    styled_fig(fig13); fig13.update_layout(yaxis=dict(autorange="reversed")); st.plotly_chart(fig13, use_container_width=True)
                else:
                    st.info("Add vendor names to see this chart.")
        with c2:
            if "product_description" in df.columns:
                pd2 = df[df["product_description"].astype(str).str.strip() != ""].copy()
                if not pd2.empty:
                    tm = (pd2.groupby("product_description").agg(margin=("margin","mean"), revenue=("selling_price","sum")).nlargest(10,"margin").reset_index())
                    tm["product_description"] = tm["product_description"].str[:30]
                    fig14 = px.bar(tm, x="margin", y="product_description", orientation="h", title="Top Products by Margin %", color="margin", color_continuous_scale=[[0,"#070C18"],[1,"#7ADFA0"]])
                    styled_fig(fig14); fig14.update_layout(yaxis=dict(autorange="reversed")); st.plotly_chart(fig14, use_container_width=True)
                else:
                    st.info("Add product descriptions to see this chart.")


def page_reminders():
    page_header("Reminders", "Payment Follow-ups")
    df = fetch_all()
    if df.empty:
        st.info("No data available.")
        return

    today_ts  = pd.Timestamp(date.today())
    df["days_old"] = (today_ts - df["sale_date"]).dt.days
    overdue_count = len(df[(df["pending_amount"] > 0) & (df["days_old"] > 30)])
    flagged_count = int((df["delay_status"] == 1).sum())

    if overdue_count or flagged_count:
        bc = st.columns(2)
        if overdue_count: bc[0].error(f"{overdue_count} payments overdue (30+ days)")
        if flagged_count: bc[1].warning(f"{flagged_count} transactions flagged")
    else:
        st.success("All clear — no overdue or flagged payments.")

    rule_sm()
    t1, t2, t3, t4 = st.tabs(["Overdue (30d+)","Flagged","High Value","Upcoming"])

    with t1:
        ov = df[(df["pending_amount"] > 0) & (df["days_old"] > 30)].sort_values("days_old", ascending=False)
        if ov.empty:
            st.success("No overdue payments.")
        else:
            st.warning(f"{len(ov)} overdue — ₹{ov['pending_amount'].sum():,.0f} total")
            for _, r in ov.iterrows():
                with st.expander(f"{r['customer_name']}  ·  ₹{r['pending_amount']:,.0f}  ·  {int(r['days_old'])} days"):
                    row_id = int(r["id"])
                    ca, cb, cc, cd = st.columns([2,2,1,1])
                    ca.write(r["sale_date"].strftime("%d %b %Y"))
                    cb.write(r.get("product_category","—"))
                    with cc:
                        if st.button("Mark Paid", key=f"op_{row_id}", use_container_width=True):
                            st.session_state.overdue_payment_editor_id = row_id
                            st.rerun()
                    with cd:
                        if st.button("Remind", key=f"or_{row_id}", use_container_width=True):
                            st.toast(f"Reminder noted for {r['customer_name']}.")
                    if st.session_state.get("overdue_payment_editor_id") == row_id:
                        st.markdown("<div class='pay-form-note'>Enter full or partial payment details.</div>", unsafe_allow_html=True)
                        with st.form(f"overdue_payment_form_{row_id}"):
                            pc1, pc2, pc3, pc4 = st.columns([1, 1, 1, 1.2])
                            with pc1:
                                _, payment_amount, payment_ok = currency_input("Amount Received (₹)", f"overdue_payment_amount_{row_id}", money_value(r.get("pending_amount")))
                            with pc2:
                                payment_date = st.date_input("Paid Date", value=date.today(), key=f"overdue_payment_date_{row_id}")
                            with pc3:
                                payment_method = st.selectbox("Paid By", PAYMENT_COLLECTION_METHODS, key=f"overdue_payment_method_{row_id}")
                            with pc4:
                                received_by = st.text_input("Received By", value=default_receiver_name(), key=f"overdue_payment_received_by_{row_id}")
                            save_col, cancel_col = st.columns(2)
                            with save_col:
                                save_payment = st.form_submit_button("Save Payment", use_container_width=True)
                            with cancel_col:
                                cancel_payment = st.form_submit_button("Cancel", use_container_width=True)
                            if save_payment:
                                if not payment_ok:
                                    st.error("Payment amount must be a valid number.")
                                else:
                                    ok, message = record_payment(r, payment_amount, payment_date, payment_method, received_by)
                                    if ok:
                                        st.session_state.overdue_payment_editor_id = None
                                        st.success(message)
                                        st.rerun()
                                    else:
                                        st.error(message)
                            if cancel_payment:
                                st.session_state.overdue_payment_editor_id = None
                                st.rerun()

    with t2:
        dl = df[df["delay_status"] == 1].sort_values("pending_amount", ascending=False)
        if dl.empty:
            st.success("No flagged payments.")
        else:
            st.error(f"{len(dl)} flagged — ₹{dl['pending_amount'].sum():,.0f}")
            show = dl[["customer_name","sale_date","product_category","selling_price","pending_amount","days_old"]].copy()
            show["sale_date"] = show["sale_date"].dt.strftime("%d %b %Y")
            show.columns = ["Customer","Date","Category","Amount ₹","Pending ₹","Days Old"]
            st.dataframe(show, use_container_width=True, hide_index=True)
            sc = st.selectbox("Clear flag for:", dl["id"].tolist(), format_func=lambda x: f"#{x} — {dl[dl['id']==x]['customer_name'].values[0]}")
            if st.button("Clear Flag"):
                get_col().update_one({"id": sc}, {"$set": {"delay_status":0}})
                invalidate_cache(); st.success("Flag cleared."); st.rerun()

    with t3:
        hv = df[df["selling_price"] >= 10000].sort_values("selling_price", ascending=False).head(20).copy()
        if hv.empty:
            st.info("No high-value sales (₹10,000+) yet.")
        else:
            hv["sale_date"]        = hv["sale_date"].dt.strftime("%d %b %Y")
            hv["payment_received"] = hv["payment_received"].map({0:"Pending",1:"Paid"})
            show = hv[["customer_name","sale_date","product_category","selling_price","profit","payment_received"]].copy()
            show.columns = ["Customer","Date","Category","Amount ₹","Profit ₹","Status"]
            st.dataframe(show, use_container_width=True, hide_index=True)

    with t4:
        soon = df[(df["pending_amount"] > 0) & (df["days_old"] >= 7) & (df["days_old"] <= 30) & (df["delay_status"] == 0)].sort_values("days_old", ascending=False)
        if soon.empty:
            st.info("No follow-ups needed in the 7–30 day window.")
        else:
            st.info(f"{len(soon)} sales with pending payments between 7–30 days old.")
            show = soon[["customer_name","customer_phone","sale_date","product_category","pending_amount","days_old"]].copy()
            show["sale_date"] = show["sale_date"].dt.strftime("%d %b %Y")
            show.columns = ["Customer","Phone","Date","Category","Pending ₹","Days Old"]
            st.dataframe(show, use_container_width=True, hide_index=True)


def page_inventory():
    page_header("Inventory", "Stock Management")
    inv_col = get_db()["inventory"]
    t1, t2 = st.tabs(["Current Stock","Add / Update Stock"])

    with t1:
        items = list(inv_col.find({}, {"_id":0}))
        if not items:
            st.markdown("<div class='empty'><div class='empty-glyph'>◆</div><div>No inventory items yet.</div></div>", unsafe_allow_html=True)
        else:
            inv_df = pd.DataFrame(items)
            total_value  = (inv_df.get("quantity", pd.Series([0])) * inv_df.get("cost_price", pd.Series([0]))).sum()
            low_stock    = inv_df[inv_df.get("quantity", pd.Series([0])) <= inv_df.get("min_stock", pd.Series([5]))]
            out_of_stock = inv_df[inv_df.get("quantity", pd.Series([0])) == 0]
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total SKUs",      len(inv_df))
            m2.metric("Inventory Value", f"₹{total_value:,.0f}")
            m3.metric("Low Stock",       len(low_stock))
            m4.metric("Out of Stock",    len(out_of_stock))
            if not low_stock.empty: st.warning(f"{len(low_stock)} item(s) running low.")
            rule_sm()
            cat_f = st.selectbox("Filter by Category", ["All"] + CATEGORIES)
            view  = inv_df.copy()
            if cat_f != "All" and "category" in view.columns: view = view[view["category"] == cat_f]
            if "quantity" in view.columns and "min_stock" in view.columns:
                view["Status"] = view.apply(lambda r: "Out of Stock" if r["quantity"]==0 else ("Low Stock" if r["quantity"]<=r["min_stock"] else "OK"), axis=1)
            st.dataframe(view, use_container_width=True, hide_index=True)
            if "category" in inv_df.columns and "quantity" in inv_df.columns:
                cat_stock = inv_df.groupby("category")["quantity"].sum().reset_index()
                fig = px.bar(cat_stock, x="category", y="quantity", title="Stock by Category", color="quantity", color_continuous_scale=[[0,"#C05060"],[0.4,"#2E6FD8"],[1,"#7ADFA0"]])
                styled_fig(fig, 260); st.plotly_chart(fig, use_container_width=True)

    with t2:
        sec("Add or Update Stock Item")
        with st.form("inv_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                item_name = st.text_input("Item Name *", placeholder="e.g. Banarasi Silk Saree")
                item_sku  = st.text_input("SKU / Code",  placeholder="e.g. SAR-001")
                item_cat  = st.selectbox("Category", CATEGORIES)
                item_vend = st.text_input("Vendor")
            with c2:
                item_qty  = st.number_input("Quantity *",       min_value=0, step=1)
                item_min  = st.number_input("Min Stock Alert",  min_value=0, step=1, value=5)
                item_cost = st.number_input("Cost Price (₹) *", min_value=0.0, step=50.0, format="%.2f")
                item_mrp  = st.number_input("Selling Price (₹)",min_value=0.0, step=50.0, format="%.2f")
            item_notes = st.text_area("Notes", height=55)
            if st.form_submit_button("Save Item", use_container_width=True):
                if not item_name.strip():
                    st.error("Item name is required.")
                else:
                    inv_col.update_one(
                        {"sku": item_sku.strip() or item_name.strip()},
                        {"$set": {"name":item_name.strip(),"sku":item_sku.strip(),"category":item_cat,"vendor":item_vend.strip(),"quantity":item_qty,"min_stock":item_min,"cost_price":round(item_cost,2),"sell_price":round(item_mrp,2),"notes":item_notes.strip(),"updated_at":str(datetime.now())}},
                        upsert=True,
                    )
                    st.success(f"'{item_name.strip()}' saved to inventory.")
                    st.rerun()

# =====================================================
# BACKUP & RESTORE PAGE
# =====================================================

def page_backup_restore():
    # ── Header ────────────────────────────────────────────────────────────
    st.markdown("""
    <div class='bk-header'>
        <span class='bk-header-icon'>🗄️</span>
        <div class='bk-header-title'>Backup &amp; Restore</div>
        <div class='bk-header-sub'>Database Checkpoint Management · Vinay Boutique</div>
    </div>
    """, unsafe_allow_html=True)

    # ── Last backup timestamp from session state ──────────────────────────
    last_backup_ts = st.session_state.get("last_backup_ts", None)
    last_restore_ts = st.session_state.get("last_restore_ts", None)

    # ══════════════════════════════════════════════════════════
    # ROW 1 — Backup & Restore cards side by side
    # ══════════════════════════════════════════════════════════
    col_bk, col_re = st.columns(2, gap="large")

    # ── Backup card ───────────────────────────────────────────────────────
    with col_bk:
        st.markdown("""
        <div class='bk-card' style='animation-delay:0s'>
            <div class='bk-card-icon bk-icon-blue'>💾</div>
            <div class='bk-card-title'>Backup Database</div>
            <div class='bk-card-desc'>
                Export a complete checkpoint of all sales, customers, and inventory data.
                Download as Excel for safekeeping or migration.
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        df = fetch_all()

        # CSV checkpoint
        csv_data = df.assign(sale_date=df["sale_date"].astype(str)).to_csv(index=False) if not df.empty else "No data"
        st.download_button(
            label="⬇️  Download CSV Checkpoint",
            data=csv_data,
            file_name=f"boutique_checkpoint_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            use_container_width=True,
        )
        st.caption("Includes all sales, customer, and payment records")

        st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)

        # Excel checkpoint
        if not df.empty:
            excel_data = to_excel(df)
            ts_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            st.download_button(
                label="⬇️  Download Excel Checkpoint",
                data=excel_data,
                file_name=f"boutique_checkpoint_{ts_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.caption("Formatted spreadsheet with column headers and styling")
        else:
            st.info("No data to export yet.")

        if last_backup_ts:
            st.markdown(f"<div class='bk-ts'>Last export: {last_backup_ts}</div>", unsafe_allow_html=True)

        if st.button("📋  Record Manual Backup Note", use_container_width=True):
            ts = datetime.now().strftime("%d %b %Y, %I:%M %p")
            st.session_state.last_backup_ts = ts
            st.success(f"✓ Manual backup noted at {ts}")
            st.rerun()

    # ── Restore card ──────────────────────────────────────────────────────
    with col_re:
        st.markdown("""
        <div class='bk-card' style='animation-delay:0.1s'>
            <div class='bk-card-icon bk-icon-green'>♻️</div>
            <div class='bk-card-title'>Restore from Checkpoint</div>
            <div class='bk-card-desc'>
                Upload a previously exported CSV checkpoint to restore records.
                Existing data will be preserved — only new records are added.
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        uploaded = st.file_uploader(
            "Upload Checkpoint File",
            type=["csv"],
            help="Upload a CSV file previously exported from this application",
            label_visibility="visible",
        )
        st.caption("200 MB max · CSV format only")

        if last_restore_ts:
            st.markdown(
                f"<span class='bk-status-badge bk-status-ok'>✓ Last restored: {last_restore_ts}</span>",
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)

        if uploaded is not None:
            try:
                restore_df = pd.read_csv(uploaded)
                row_count = len(restore_df)
                col_count = len(restore_df.columns)

                st.markdown(f"""
                <div class='bk-card' style='animation-delay:0.15s; border-color: rgba(37,99,235,0.3); margin-top:0.8rem'>
                    <div class='bk-card-title' style='font-size:0.85rem'>📊 File Preview</div>
                    <div class='bk-card-desc'>
                        <b>{row_count}</b> records · <b>{col_count}</b> columns detected<br>
                        Columns: {', '.join(restore_df.columns[:6].tolist())}{' …' if col_count > 6 else ''}
                    </div>
                </div>
                """, unsafe_allow_html=True)

                st.dataframe(restore_df.head(5), use_container_width=True, hide_index=True)

                st.warning(
                    "⚠️  This will **insert** records from the checkpoint into the live database. "
                    "Duplicate entries may result if records already exist.",
                )

                confirm = st.checkbox("I understand — proceed with restore")
                if confirm:
                    if st.button("🔄  Restore Database from Checkpoint", use_container_width=True):
                        progress_placeholder = st.empty()
                        progress_placeholder.markdown(
                            "<div style='background:var(--bg-2);border-radius:999px;overflow:hidden;height:6px;margin:0.5rem 0'>"
                            "<div class='bk-progress-bar'></div></div>",
                            unsafe_allow_html=True,
                        )
                        status_msg = st.empty()
                        status_msg.info("Restoring records…")

                        # ── Actual restore logic ──────────────────────────
                        inserted = 0
                        skipped  = 0
                        errors   = []

                        # ── Column name normalisation ─────────────────────
                        # The app exports title-cased columns (e.g. "Customer Name")
                        # but internally uses snake_case (e.g. "customer_name").
                        # Build a mapping: title-cased export name → internal name.
                        EXPORT_TO_INTERNAL = {
                            "Id":                  "id",
                            "Customer Name":       "customer_name",
                            "Customer Phone":      "customer_phone",
                            "Sale Date":           "sale_date",
                            "Product Category":    "product_category",
                            "Product Description": "product_description",
                            "Vendor":              "vendor",
                            "Buying Price":        "buying_price",
                            "Selling Price":       "selling_price",
                            "Profit":              "profit",
                            "Profit Margin %":     "margin",
                            "Amount Paid":         "amount_paid",
                            "Pending Amount":      "pending_amount",
                            "Payment Status":      "_payment_status_str",  # converted below
                            "Status":              "_payment_status_str",
                            "Delayed":             "_delayed_str",          # converted below
                            "Payment Method":      "payment_method",
                            "Notes":               "notes",
                            "Created At":          "created_at",
                        }
                        # Rename columns that match the export format; leave unknown ones as-is
                        restore_df = restore_df.rename(
                            columns={k: v for k, v in EXPORT_TO_INTERNAL.items() if k in restore_df.columns}
                        )
                        # Also lowercase any remaining columns that weren't renamed
                        restore_df.columns = [
                            c.lower().replace(" ", "_") if c not in restore_df.columns else c
                            for c in restore_df.columns
                        ]

                        required_cols = {"customer_name", "selling_price"}
                        if not required_cols.issubset(set(restore_df.columns)):
                            progress_placeholder.empty()
                            status_msg.empty()
                            st.error(
                                f"Invalid checkpoint file. Required columns missing: "
                                f"{required_cols - set(restore_df.columns)}"
                            )
                        else:
                            for _, row in restore_df.iterrows():
                                try:
                                    doc = row.dropna().to_dict()

                                    # Convert "Payment Status" string → payment_received int
                                    if "_payment_status_str" in doc:
                                        ps = str(doc.pop("_payment_status_str")).strip().lower()
                                        doc["payment_received"] = 1 if ps in ("paid", "received", "1") else 0

                                    # Convert "Delayed" string → delay_status int
                                    if "_delayed_str" in doc:
                                        dl = str(doc.pop("_delayed_str")).strip().lower()
                                        doc["delay_status"] = 1 if dl in ("yes", "true", "1") else 0

                                    # Derive payment_received from pending_amount if not set
                                    if "payment_received" not in doc:
                                        pending = float(doc.get("pending_amount", 0) or 0)
                                        doc["payment_received"] = 0 if pending > 0 else 1

                                    # Default delay_status
                                    if "delay_status" not in doc:
                                        doc["delay_status"] = 0

                                    # Normalise numeric types
                                    for num_col in ["buying_price", "selling_price", "amount_paid", "pending_amount", "quantity", "profit", "margin"]:
                                        if num_col in doc:
                                            try:
                                                doc[num_col] = float(doc[num_col])
                                            except (ValueError, TypeError):
                                                doc.pop(num_col, None)
                                    for int_col in ["payment_received", "delay_status"]:
                                        if int_col in doc:
                                            doc[int_col] = int(doc[int_col])

                                    # Drop the old exported id — assign a fresh one
                                    doc.pop("id", None)
                                    doc["id"]          = get_next_id()
                                    doc["restored_at"] = str(datetime.now())
                                    get_col().insert_one(doc)
                                    inserted += 1
                                except Exception as e:
                                    skipped += 1
                                    errors.append(str(e))

                            invalidate_cache()
                            progress_placeholder.empty()
                            status_msg.empty()
                            ts = datetime.now().strftime("%d %b %Y, %I:%M %p")
                            st.session_state.last_restore_ts = ts

                            if inserted > 0:
                                st.success(
                                    f"✅  Restore complete — **{inserted}** records inserted"
                                    + (f", {skipped} skipped." if skipped else ".")
                                )
                            if errors:
                                with st.expander(f"⚠️  {len(errors)} row(s) had errors"):
                                    for err in errors[:10]:
                                        st.caption(err)
                            st.rerun()

            except Exception as e:
                st.error(f"Could not read file: {e}")

    # ══════════════════════════════════════════════════════════
    # ROW 2 — Database stats card
    # ══════════════════════════════════════════════════════════
    rule()
    st.markdown("""
    <div class='bk-card' style='animation-delay:0.2s'>
        <div class='bk-card-title'>📈  Current Database Status</div>
        <div class='bk-card-desc'>Live snapshot of records in the database</div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    df2 = fetch_all()
    m = metrics(df2)
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total Records",  m["sales"])
    c2.metric("Total Revenue",  f"₹{m['revenue']:,.0f}")
    c3.metric("Unique Customers", m["customers"])
    c4.metric("Pending Payments", f"₹{m['pending']:,.0f}")
    c5.metric("Data Health",    "✓ Live" if m["sales"] > 0 else "Empty")

    rule_sm()
    st.caption(f"Database last queried: {datetime.now().strftime('%d %b %Y, %I:%M:%S %p')}  ·  Boutique Manager v2.0")


# =====================================================
# MAIN
# =====================================================

def main():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    if "theme" not in st.session_state:
        st.session_state.theme = "light"

    # Apply light mode CSS overrides if needed
    inject_theme()

    if not st.session_state.logged_in:
        page_add_sale(public=True)
        render_admin_login_strip()
        return

    page = sidebar()

    if   "Dashboard"   in page: page_dashboard()
    elif "Add Sale"    in page: page_add_sale(public=False)
    elif "Review"      in page: page_review()
    elif "Update"      in page: page_update()
    elif "Customer"    in page: page_customers()
    elif "Analytics"   in page: page_analytics()
    elif "Reminders"   in page: page_reminders()
    elif "Backup"      in page: page_backup_restore()
    elif "Logout"      in page:
        st.session_state.logged_in = False
        st.session_state.username  = None
        st.rerun()

if __name__ == "__main__":
    main()
